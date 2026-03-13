import os  
from flask import Flask,render_template, request, jsonify, redirect, url_for, session ,send_from_directory
import mysql.connector
from mysql.connector import pooling
from flask_session import Session
import pymysql.cursors 
import cv2
import numpy as np
import io
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import Workbook, load_workbook
from flask import send_file
import json
from datetime import datetime

app = Flask(__name__)  
app.secret_key = 'your_secret_key'  

UPLOAD_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "AssessmentPictures")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

UPLOAD_IMG = os.path.join(os.path.expanduser("~"), "Desktop", "RemedyPictures")
app.config["UPLOAD_IMG"] = UPLOAD_IMG

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

if not os.path.exists(UPLOAD_IMG):
    os.makedirs(UPLOAD_IMG)

# Connect to MySQL
#def get_db_connection():
#    return mysql.connector.connect(
#        host="localhost",  # Connecting to localhost
#        user="root",
#        password="nare@2058",
#        database="remedydb",
#        autocommit=True,  # Prevents timeout issues
#        connection_timeout=300000  # Keeps connection alive
# )

db_pool = pooling.MySQLConnectionPool(
    pool_name="mypool",
    pool_size=10,  # Adjust based on workload
    host="localhost",  # Or server IP if hosted remotely
    user="root",
    password="nare@2058", 
    database="remedydb",
    autocommit=True,  # Prevents timeout issues
    connection_timeout=60 # 60 seconds is ideal
   
)
def get_db_connection():
    return db_pool.get_connection()

import pandas as pd
@app.route("/update_assessment_rem_case", methods=["POST"])
def update_assessment_rem_case():
    connection = None
    cursor = None

    try:
        data = request.json
        table_id = data["table_id"]
        pile_no = data["pile_no"]
        assessment_case = data["assessment_case"]

        connection = get_db_connection()
        cursor = connection.cursor()

        cursor.execute("""
            UPDATE assessment
            SET `Assessment case` = %s
            WHERE `Table ID` = %s
              AND `Pile No` = %s
        """, (assessment_case, table_id, pile_no))

        connection.commit()  # optional since autocommit=True

        return jsonify({"success": True})

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
@app.route('/get_case_data')
def get_case_data():

    area_id = request.args.get('area_id')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        query = """
            SELECT 
                `Table ID` AS table_id,
                `Pile No` AS pile_no,
                `Assessment case` AS assessment_case
            FROM assessment
            WHERE `Area ID` = %s
        """

        cursor.execute(query, (area_id,))
        rows = cursor.fetchall()

        if not rows:
            return jsonify({
                "success": False,
                "message": "No case data found",
                "data": []
            })

        return jsonify({
            "success": True,
            "data": rows
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

    finally:
        cursor.close()
        connection.close()
@app.route('/download_tableinfo')
def download_tableinfo():
    area_id = request.args.get('area_id')
    user_id = request.args.get('user_id')
    task_date = request.args.get('task_date')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        query = """
            SELECT `Assessment ID`, `Table ID`, `Pile No`
            FROM assessment
            WHERE `Area ID` = %s AND `User ID` = %s AND `Task Date` = %s
        """
        cursor.execute(query, (area_id, user_id, task_date))
        rows = cursor.fetchall()

        if not rows:
            return jsonify({
                'success': False,
                'message': 'No records found for the selected filters.'
            }), 404

        # Convert to Excel
        df = pd.DataFrame(rows)

        # ✅ 1. Save locally on your system
        save_path = r"C:\Users\LENOVO\Desktop\nocoderename.xlsx"
        df.to_excel(save_path, index=False, sheet_name='TableInfo')

        # ✅ 2. Also send the same file to browser for download
        return send_file(
            save_path,
            as_attachment=True,
            download_name="nocoderename.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    finally:
        cursor.close()
        connection.close()
@app.route('/get_images')
def get_images():
    data = []

    base = app.config["CONCRETE_FOLDER"]

    # 🔥 Sort folders by creation time (like created order)
    folders = sorted(
        os.listdir(base),
        key=lambda x: os.path.getctime(os.path.join(base, x))
    )

    for folder in folders:
        folder_path = os.path.join(base, folder)

        if os.path.isdir(folder_path):

            images = sorted(
                [f for f in os.listdir(folder_path)
                 if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))],
                key=lambda x: os.path.getctime(os.path.join(folder_path, x))
            )

            data.append({
                "folder": folder,
                "images": images
            })

    return render_template('concreteimgrep.html', data=data)



@app.route('/concrete_image/<folder>/<filename>')
def concrete_image(folder, filename):
    return send_from_directory(
        os.path.join(app.config["CONCRETE_FOLDER"], folder),
        filename
    )
CONCRETE_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "Concreteprocess")
app.config["CONCRETE_FOLDER"] = CONCRETE_FOLDER
if not os.path.exists(CONCRETE_FOLDER):
    os.makedirs(CONCRETE_FOLDER)
@app.route('/download_concrete_pdf', methods=['POST'])
def save_concrete_pdf():
    pdf_file = request.files['pdf']

    # Get Desktop Path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Create folder: Desktop/Concrete Reports/Concrete Process
    reports_folder = os.path.join(desktop_path, "Concrete Reports", "Concrete Process")
    os.makedirs(reports_folder, exist_ok=True)

    # Extract filename
    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)

    # Avoid duplicate names
    pdf_path = os.path.join(reports_folder, original_filename)
    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(
            reports_folder,
            f"{base_name}_{count}{extension}"
        )
        count += 1

    # Save PDF
    pdf_file.save(pdf_path)

    return jsonify({
        "message": f"Concrete Process PDF saved successfully as {os.path.basename(pdf_path)}!"
    })

@app.route('/download_rusttableinfo')
def download_rusttableinfo():
    area_id = request.args.get('area_id')
    user_id = request.args.get('user_id')
    task_date = request.args.get('task_date')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        query = """
            SELECT `Assessment ID`, `Table ID`, `Pile No`
            FROM bracingrust
            WHERE `Area ID` = %s AND `User ID` = %s AND `Task Date` = %s
        """
        cursor.execute(query, (area_id, user_id, task_date))
        rows = cursor.fetchall()

        if not rows:
            return jsonify({
                'success': False,
                'message': 'No records found for the selected filters.'
            }), 404

        # Convert to Excel
        df = pd.DataFrame(rows)

        # ✅ 1. Save locally on your system
        save_path = r"C:\Users\LENOVO\Desktop\rustids.xlsx"
        df.to_excel(save_path, index=False, sheet_name='TableInfo')

        # ✅ 2. Also send the same file to browser for download
        return send_file(
            save_path,
            as_attachment=True,
            download_name="rustids.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_work', methods=['POST'])
def submit_work():
    data = request.form
    work_type = data.get('work_type')

    connection = get_db_connection()
    cursor = connection.cursor()

    if work_type == '30mm Excavation':
        sql = """
            INSERT INTO excavation_30mm
            (`Area ID`, `Contractor ID`, `Sub Contractor ID`,
             `Selected Hotspots`, `Task Date`, `30mm Excavation`)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        values = (
            data.get('area_id'),
            data.get('contractor_id'),
            data.get('sub_contractor_id'),
            data.get('selectedHotspots'),
            data.get('task_date'),
            'Completed'
        )

    elif work_type == '75mm Excavation':
        sql = """
            INSERT INTO excavation_75mm
            (`Area ID`, `Contractor ID`, `Sub Contractor ID`,
             `Selected Hotspots`, `Task Date`, `75mm Excavation`)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        values = (
            data.get('area_id'),
            data.get('contractor_id'),
            data.get('sub_contractor_id'),
            data.get('selectedHotspots'),
            data.get('task_date'),
            'Completed'
        )

    elif work_type == 'Concreted':
        sql = """
            INSERT INTO concreted_work
            (`Area ID`, `Contractor ID`, `Sub Contractor ID`,
             `Selected Hotspots`, `Task Date`, `Concreted`)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        values = (
            data.get('area_id'),
            data.get('contractor_id'),
            data.get('sub_contractor_id'),
            data.get('selectedHotspots'),
            data.get('task_date'),
            'Completed'
        )
    else:
        return jsonify({"success": False, "message": "Invalid Work Type"}), 400

    cursor.execute(sql, values)
    connection.commit()

    # ✅ GET AUTO-INCREMENT ID
    numeric_work_id = cursor.lastrowid
    display_work_id = f"W{numeric_work_id}"

    cursor.close()
    connection.close()

    return jsonify({
        "success": True,
        "message": "Work data saved successfully",
        "work_id": display_work_id   # 🔥 W1, W2, W3
    })
@app.route('/get_work_info')
def get_work_info():
    area_id = request.args.get('area_id')
    contractor_id = request.args.get('contractor_id')
    task_date = request.args.get('task_date')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    tables = [
        ("30mm Excavation", "excavation_30mm"),
        ("75mm Excavation", "excavation_75mm"),
        ("Concreted", "concreted_work")
    ]

    result = []

    for work_type, table in tables:
        query = f"""
            SELECT
                '{work_type}' AS work_type,
                t.`Area ID` AS area_id,
                u.`User Name` AS contractor_name,
                t.`Task Date` AS task_date,
                IF(t.`Selected Hotspots` IS NULL OR t.`Selected Hotspots`='',
                    0,
                    LENGTH(t.`Selected Hotspots`)
                    - LENGTH(REPLACE(t.`Selected Hotspots`, ',', '')) + 1
                ) AS hotspot_count
            FROM `{table}` t
            JOIN users u
              ON u.`User ID` = t.`Contractor ID`
            WHERE t.`Area ID` = %s
              AND t.`Contractor ID` = %s
        """

        params = [area_id, contractor_id]

        if task_date:
            query += " AND t.`Task Date` = %s"
            params.append(task_date)

        cursor.execute(query, params)
        result.extend(cursor.fetchall())

    cursor.close()
    connection.close()

    return jsonify({"success": True, "data": result})


@app.route('/download_work_report')
def download_work_report():
    area_id = request.args.get('area_id')
    contractor_id = request.args.get('contractor_id')
    task_date = request.args.get('task_date')

    connection = get_db_connection()
    cursor = connection.cursor()

    queries = [
        ("30mm Excavation", "excavation_30mm"),
        ("75mm Excavation", "excavation_75mm"),
        ("Concreted", "concreted_work")
    ]

    rows = []
    rows.append(["Work Type", "Area ID", "Contractor ID", "Task Date", "Hotspot Count"])

    for work_type, table in queries:
        query = f"""
            SELECT
                `Area ID`,
                `Contractor ID`,
                `Task Date`,
                IF(`Selected Hotspots` IS NULL OR `Selected Hotspots`='',
                    0,
                    LENGTH(`Selected Hotspots`) - LENGTH(REPLACE(`Selected Hotspots`, ',', '')) + 1
                )
            FROM `{table}`
            WHERE `Area ID`=%s AND `Contractor ID`=%s
        """

        params = [area_id, contractor_id]

        if task_date:
            query += " AND `Task Date`=%s"
            params.append(task_date)

        cursor.execute(query, params)
        for r in cursor.fetchall():
            rows.append([work_type, *r])

    cursor.close()
    connection.close()

    import csv
    from flask import Response

    def generate():
        for row in rows:
            yield ",".join(map(str, row)) + "\n"

    return Response(
        generate(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=work_report.csv"}
    )

@app.route('/download_remedytableinfo')
def download_remedytableinfo():
    area_id = request.args.get('area_id')
    user_id = request.args.get('user_id')
    task_date = request.args.get('task_date')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        query = """
            SELECT `Remedy ID`, `Table ID`, `Pile No`
            FROM remedy
            WHERE `Area ID` = %s AND `User ID` = %s AND `Task Date` = %s
        """
        cursor.execute(query, (area_id, user_id, task_date))
        rows = cursor.fetchall()

        if not rows:
            return jsonify({
                'success': False,
                'message': 'No records found for the selected filters.'
            }), 404

        # Convert to Excel
        df = pd.DataFrame(rows)

        # ✅ 1. Save locally on your system
        save_path = r"C:\Users\LENOVO\Desktop\remedyfolderids.xlsx"
        df.to_excel(save_path, index=False, sheet_name='TableInfo')

        # ✅ 2. Also send the same file to browser for download
        return send_file(
            save_path,
            as_attachment=True,
            download_name="remedyfolderids.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    finally:
        cursor.close()
        connection.close()

def notify_cloud_server(change_type, table, data):
    payload = {
        "change_type": change_type,
        "table": table,
        "data": data
    }

    try:
        response = requests.post("http://89.116.122.75:5000/api/sync", json=payload)
        print("Cloud sync response:", response.status_code, response.text)
    except Exception as e:
        print("Sync failed:", e)
@app.route('/get-assessment-count', methods=['POST'])
def get_assessment_count():
    try:
        data = request.get_json()
        selected_date = data.get("date")

        connection = get_db_connection()
        cursor = connection.cursor()

        # Total count
        query_total = """
            SELECT COUNT(*) FROM assessment
            WHERE DATE(`Task Date`) = %s
        """
        cursor.execute(query_total, (selected_date,))
        total_count = cursor.fetchone()[0]

        # Case-wise count
        query_cases = """
            SELECT `Assessment case`, COUNT(*) 
            FROM assessment 
            WHERE DATE(`Task Date`) = %s 
            GROUP BY `Assessment case`
        """
        cursor.execute(query_cases, (selected_date,))
        case_counts = {row[0]: row[1] for row in cursor.fetchall()}

        return jsonify({
            "success": True,
            "count": total_count,
            "case1": case_counts.get("Case1", 0),
            "case2": case_counts.get("Case2", 0),
            "case3": case_counts.get("Case3", 0),
            "case4": case_counts.get("Case4", 0)
        })

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        cursor.close()
        connection.close()
@app.route('/get-remedy-count', methods=['POST'])
def get_remedy_count():
    try:
        data = request.get_json()
        selected_date = data.get("date")

        connection = get_db_connection()
        cursor = connection.cursor()

        # Total count
        query_total = """
            SELECT COUNT(*) FROM remedy
            WHERE DATE(`Task Date`) = %s
        """
        cursor.execute(query_total, (selected_date,))
        total_count = cursor.fetchone()[0]

        # Case-wise count
        query_cases = """
            SELECT `Assessed Case`, COUNT(*) 
            FROM remedy
            WHERE DATE(`Task Date`) = %s 
            GROUP BY `Assessed Case`
        """
        cursor.execute(query_cases, (selected_date,))
        case_counts = {row[0]: row[1] for row in cursor.fetchall()}

        return jsonify({
            "success": True,
            "count": total_count,
            "case1": case_counts.get("Case1", 0),
            "case2": case_counts.get("Case2", 0),
            "case3": case_counts.get("Case3", 0),
            "case4": case_counts.get("Case4", 0)
        })

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        cursor.close()
        connection.close()


@app.route('/chart-data')
def get_chart_data():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Monthly Assessment Count based on "Task Date"
        cursor.execute("""
            SELECT MONTHNAME(`Task Date`) AS month, COUNT(*) AS count
            FROM assessment
            GROUP BY MONTH(`Task Date`)
            ORDER BY MONTH(`Task Date`)
        """)
        month_data = cursor.fetchall()

        # Case Count based on "Assessment case"
        cursor.execute("""
            SELECT `Assessment case` AS case_type, COUNT(*) AS count
            FROM assessment
            WHERE `Task Date` >= '2025-10-01'
            GROUP BY `Assessment case`
        """)
        case_data = cursor.fetchall()

        # Replace "Not Assessed" with "Not Approved"
        for case in case_data:
            if case['case_type'] == 'Not Assessed':
                case['case_type'] = 'Not Approved'

        return jsonify({
            "monthData": month_data,
            "caseData": case_data
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": "An error occurred",
            "error": str(e)
        })

    finally:
        cursor.close()
        connection.close()

@app.route('/remedy-chart-data')
def get_remedy_chart_data():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Monthly Remedy Count based on "Task Date"
        cursor.execute("""
            SELECT MONTHNAME(`Task Date`) AS month, COUNT(*) AS count
            FROM remedy
            GROUP BY MONTH(`Task Date`)
            ORDER BY MONTH(`Task Date`)
        """)
        month_data = cursor.fetchall()

        # # Case Count based on "Assessed Case"
        # cursor.execute("""
        #     SELECT Assessed Case AS case_type, COUNT(*) AS count
        #     FROM remedy
        #     GROUP BY Assessed Case
        # """)
        # case_data = cursor.fetchall()

        # # Replace "Not Assessed" with "Not Approved"
        # for case in case_data:
        #     if case['case_type'] == 'Not Assessed':
        #         case['case_type'] = 'Not Approved'
        # Case Count from assessment table using matching Table ID & Pile No
        cursor.execute("""
            SELECT 
                a.`Assessment case` AS case_type, 
                COUNT(*) AS count
            FROM remedy r
            JOIN assessment a 
              ON r.`Table ID` = a.`Table ID`
             AND r.`Pile No` = a.`Pile No`
            GROUP BY a.`Assessment case`
        """)
        case_data = cursor.fetchall()

        # Replace "Not Assessed" with "Not Approved"
        for case in case_data:
            if case['case_type'] in (None, '', 'Not Assessed'):
                case['case_type'] = 'Not Approved'

        return jsonify({
            "monthData": month_data,
            "caseData": case_data
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": "An error occurred",
            "error": str(e)
        })

    finally:
        cursor.close()
        connection.close()

# below for work info
@app.route('/get_assessment_info', methods=['POST'])
def get_assessment_info():
    data = request.get_json()
    user_id = data.get('user_id') or None
    area_id = data.get('area_id')

    start_date = data.get('start_date') or None
    end_date = data.get('end_date') or None

    if not area_id:
        return jsonify({"success": False, "message": "Area ID is required."})

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor(dictionary=True)

    try:
        query = """
            SELECT 
                DATE(`Task Date`) AS date,
                COUNT(*) AS total,

                SUM(CASE WHEN `Assessment Status` = 'In Progress' THEN 1 ELSE 0 END) AS in_progress,
                SUM(CASE WHEN `Assessment Status` = 'Completed' THEN 1 ELSE 0 END) AS completed,
                SUM(CASE WHEN `Assessment Status` = 'PM Approved' THEN 1 ELSE 0 END) AS pm_approved,
                SUM(CASE WHEN `Assessment Status` = 'OE Approved' THEN 1 ELSE 0 END) AS oe_approved,

                SUM(CASE WHEN `Assessment case` = 'Not Assessed' THEN 1 ELSE 0 END) AS not_assessed,
                SUM(CASE WHEN `Assessment case` = 'Case1' THEN 1 ELSE 0 END) AS case1,
                SUM(CASE WHEN `Assessment case` = 'Case2' THEN 1 ELSE 0 END) AS case2,
                SUM(CASE WHEN `Assessment case` = 'Case3' THEN 1 ELSE 0 END) AS case3,
                SUM(CASE WHEN `Assessment case` = 'Case4' THEN 1 ELSE 0 END) AS case4

            FROM assessment
            WHERE (%s IS NULL OR `User ID` = %s)
              AND `Area ID` = %s
              AND (%s IS NULL OR `Task Date` >= %s)
              AND (%s IS NULL OR `Task Date` <= %s)
            GROUP BY `Task Date`
            ORDER BY `Task Date` DESC

        """

        cursor.execute(query, (user_id, user_id, area_id, start_date, start_date, end_date, end_date))
        rows = cursor.fetchall()


        return jsonify({"success": True, "rows": rows})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching info: {e}"}), 500

    finally:
        cursor.close()
        connection.close()

@app.route('/get_remedy_info', methods=['POST'])
def get_remedy_info():
    data = request.get_json()

    user_id = data.get('user_id') or None
    area_id = data.get('area_id')

    start_date = data.get('start_date') or None
    end_date = data.get('end_date') or None

    if not area_id:
        return jsonify({"success": False, "message": "Area ID is required."})

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor(dictionary=True)

    try:

        query = """
            SELECT 
                DATE(r.`Task Date`) AS date,
                COUNT(*) AS total,

                -- CASE COUNTS FROM ASSESSMENT TABLE
                SUM(CASE WHEN a.`Assessment Case` = 'Not Assessed' THEN 1 ELSE 0 END) AS not_assessed,
                SUM(CASE WHEN a.`Assessment Case` = 'Case1' THEN 1 ELSE 0 END) AS case1,
                SUM(CASE WHEN a.`Assessment Case` = 'Case2' THEN 1 ELSE 0 END) AS case2,
                SUM(CASE WHEN a.`Assessment Case` = 'Case3' THEN 1 ELSE 0 END) AS case3,
                SUM(CASE WHEN a.`Assessment Case` = 'Case4' THEN 1 ELSE 0 END) AS case4

            FROM remedy r
            LEFT JOIN assessment a
                ON a.`Area ID` = r.`Area ID`
            AND a.`Table ID` = r.`Table ID`
            AND a.`Pile No` = r.`Pile No`

            WHERE (%s IS NULL OR r.`User ID` = %s)
            AND r.`Area ID` = %s
            AND (%s IS NULL OR r.`Task Date` >= %s)
            AND (%s IS NULL OR r.`Task Date` <= %s)
            AND r.`Picture Location` IS NOT NULL
            AND r.`Picture Location` <> ''

            GROUP BY DATE(r.`Task Date`)
            ORDER BY DATE(r.`Task Date`) DESC
        """

        cursor.execute(query, (
            user_id, user_id,
            area_id,
            start_date, start_date,
            end_date, end_date
        ))

        rows = cursor.fetchall()
        return jsonify({"success": True, "rows": rows})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching info: {e}"}), 500

    finally:
        cursor.close()
        connection.close()

@app.route('/submit_comment', methods=['POST'])
def submit_comment():
    data = request.get_json()
    conn = None
    cursor = None

    try:
        if 'email' not in session:
            return jsonify({"success": False, "error": "User not logged in"}), 401

        email = session['email']

        conn = get_db_connection()
        cursor = conn.cursor()

        # Get logged-in user ID for Commented By
        cursor.execute("SELECT `User ID` FROM users WHERE Email = %s", (email,))
        result = cursor.fetchone()
        if not result:
            return jsonify({"success": False, "error": "Logged-in user not found"}), 404
        commented_by = result[0]

        # Data from frontend
        selected_user_id = data["user_id"]
        area_id = data["area_id"]
        table_id = data["table_id"]
        pile_no = int(data["pile_no"])  # Store Pile No directly
        assessment_case = data["case_type"]
        assessment_status = data.get("assessment_status", "Not Provided")
        comment_desc = data["comment_description"]

        # Get Task Date from assessment table
        cursor.execute("""
            SELECT `Task Date` 
            FROM assessment
            WHERE `User ID` = %s AND `Area ID` = %s AND `Table ID` = %s AND `Pile No` = %s
            LIMIT 1
        """, (selected_user_id, area_id, table_id, pile_no))
        task_date_row = cursor.fetchone()
        if not task_date_row:
            return jsonify({"success": False, "error": "Task Date not found for this pile"}), 404
        date_posted = task_date_row[0]  # e.g., '07 Jun 2025'


        # Generate Comment ID like C00001
        cursor.execute("SELECT `Comment ID` FROM comments ORDER BY `Comment ID` DESC LIMIT 1")
        last_id_row = cursor.fetchone()
        if last_id_row:
            last_id_str = str(last_id_row[0])
            last_number = int(last_id_str[1:]) if last_id_str.startswith("C") else 0
        else:
            last_number = 0
        new_comment_id = f"C{last_number + 1:05d}"

        # Insert comment using new structure
        cursor.execute("""
            INSERT INTO comments
            (`Comment ID`, `Comment Type`, `User ID`, `Area ID`, `Table ID`, `Pile No`,
             `Date Posted`, `Comment Description`, `Comment Date`, `Commented By`,
             `Assessment Status`, `Assessment Case`)
            VALUES
            (%s, 'Assessment', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            new_comment_id, selected_user_id, area_id, table_id, pile_no, 
            date_posted, comment_desc, date_posted, commented_by,
            assessment_status, assessment_case
        ))

        conn.commit()
        return jsonify({"success": True, "comment_id": new_comment_id})

    except Exception as e:
        print("Error:", e)
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

from flask import send_from_directory

# Folder where you want to save the file (renamed as requested)
assessment_html_changes = r'C:\Users\LENOVO\Desktop\HTMLreport'
@app.route("/save_assessment_changes", methods=["POST"])
def save_assessment_changes():
    try:
        data = request.json.get("data", [])

        # Create a timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"assessment_changes_{timestamp}.json"

        # Full path
        file_path = os.path.join(assessment_html_changes, filename)

        # Ensure folder exists
        os.makedirs(assessment_html_changes, exist_ok=True)

        # Save to file
        with open(file_path, "w") as f:
            json.dump(data, f, indent=4)

        return jsonify({"message": "Assessment changes saved successfully!", "file": filename})
    
    except Exception as e:
        print("Error:", e)
        return jsonify({"message": "Error saving data."}), 500

# Set path where your images are saved
ASSESSMENT_PIC_FOLDER = r'C:\Users\LENOVO\Desktop\AssessmentPictures'

@app.route('/AssessmentPictures/<folder>/<filename>')
def serve_assessment_image(folder, filename):
    return send_from_directory(os.path.join(ASSESSMENT_PIC_FOLDER, folder), filename)

# Set path where your remedy images are saved
REMEDY_PIC_FOLDER = r'C:\Users\LENOVO\Desktop\RemedyPictures'
@app.route('/RemedyPictures/<folder>/<filename>')
def serve_remedy_image(folder, filename):
    return send_from_directory(os.path.join(REMEDY_PIC_FOLDER, folder), filename)
    
# Define the base folder path for images and case folders
BASE_FOLDER = 'static/images'  # Base folder path for images
CASE_FOLDERS = ['case1', 'case2', 'case3', 'case4']
def mask_metal_region(image):
    """Mask the background and keep only the metal pile region using color thresholding."""
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Metal piles are typically grayish, so target low saturation
    lower_metal = np.array([0, 0, 60])
    upper_metal = np.array([180, 70, 255])

    # Create mask where metal is likely present
    metal_mask = cv2.inRange(hsv, lower_metal, upper_metal)

    # Apply mask to keep only metal region
    result = cv2.bitwise_and(image, image, mask=metal_mask)

    return result

def calculate_rust_percentage(base_img, test_img):
    # Convert to HSV
    hsv_base = cv2.cvtColor(base_img, cv2.COLOR_BGR2HSV)
    hsv_test = cv2.cvtColor(test_img, cv2.COLOR_BGR2HSV)

    # Define a range for rust color in HSV (you can adjust as needed)
    lower_red = np.array([0, 50, 50])
    upper_red = np.array([10, 255, 255])

    # Create masks for rust areas
    base_mask = cv2.inRange(hsv_base, lower_red, upper_red)
    test_mask = cv2.inRange(hsv_test, lower_red, upper_red)

    # Find overlapping rust regions (base vs test)
    overlap = cv2.bitwise_and(base_mask, test_mask)
    match_pixels = np.count_nonzero(overlap)
    total_pixels = np.count_nonzero(base_mask)

    if total_pixels == 0:
        return 0.0  # Avoid division by zero if no rust in base

    match_percent = (match_pixels / total_pixels) * 100
    return match_percent
def detect_rust_and_damage_percentage(image):
    """Detect rust (red color) and damage (bare iron) percentage in the image."""
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Define red ranges in HSV for rust detection
    lower_red1 = np.array([0, 70, 50])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([160, 70, 50])
    upper_red2 = np.array([180, 255, 255])

    # Define low saturation and high value for detecting bare iron (damaged surface)
    lower_bare = np.array([0, 0, 100])  # Low saturation, high value (bare metal)
    upper_bare = np.array([180, 50, 255])

    # Mask for rust (red) and bare metal (damage)
    mask_rust1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask_rust2 = cv2.inRange(hsv, lower_red2, upper_red2)
    mask_bare = cv2.inRange(hsv, lower_bare, upper_bare)

    # Combine the masks for rust and bare metal (damage)
    rust_pixels = cv2.countNonZero(mask_rust1) + cv2.countNonZero(mask_rust2)
    damage_pixels = cv2.countNonZero(mask_bare)
    total_pixels = image.shape[0] * image.shape[1]

    rust_percentage = (rust_pixels / total_pixels) * 100
    damage_percentage = (damage_pixels / total_pixels) * 100

    return rust_percentage, damage_percentage

def get_max_values_for_case(case_folder, base_image):
    """Get the max rust and damage percentage for a given case folder."""
    max_rust = 0
    max_damage = 0
    
    # Go through each side image in the case folder (side1.jpg, side2.jpg, side3.jpg, side4.jpg)
    for side in range(1, 5):
        side_image_path = os.path.join(BASE_FOLDER, case_folder, f"side{side}.jpg")
        
        # Check if the side image exists
        if not os.path.exists(side_image_path):
            continue
        
        # Read the side image
        side_img = cv2.imread(side_image_path)
        
        # Calculate rust and damage percentages
        rust_percent, damage_percent = detect_rust_and_damage_percentage(side_img)
        
        # Update the max values if current values are higher
        max_rust = max(max_rust, rust_percent)
        max_damage = max(max_damage, damage_percent)
    
    return max_rust, max_damage

# @app.route('/analyze_corrosion', methods=['POST'])
# def analyze_corrosion():
#     uploaded_images = []
#     rust_scores_by_case = {case: [] for case in CASE_FOLDERS}

#     # Step 1: Read uploaded images (side1 to side4)
#     for i in range(1, 5):
#         file = request.files.get(f'side{i}')
#         if not file:
#             print(f"❌ Missing uploaded image: side{i}")
#             return jsonify({"error": f"Missing image: side{i}"}), 400

#         # print(f"✅ Received: {file.filename}")
#         file_bytes = np.frombuffer(file.read(), np.uint8)
#         img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
#         if img is None:
#             return jsonify({"error": f"Cannot decode image: side{i}"}), 400
#         img = mask_metal_region(img)
#         img = cv2.resize(img, (300, 300))


#         uploaded_images.append(img)

#     # Step 2: Compare each uploaded image to corresponding base image in each case
#     for case in CASE_FOLDERS:
#         # print(f"📁 Analyzing against: {case}")
#         case_scores = []
#         for i in range(1, 5):  # side1 to side4
#             base_img_path = os.path.join(BASE_FOLDER, case, f'side{i}.jpg')
#             if not os.path.exists(base_img_path):
#                 # print(f"⚠️ Missing base image: {base_img_path}")
#                 case_scores.append(0)
#                 continue

#             base_img = cv2.imread(base_img_path)
#             base_img = cv2.resize(base_img, (300, 300))

#             # Calculate rust/damage % between base and uploaded
#             rust_percent = calculate_rust_percentage(base_img, uploaded_images[i - 1])
#             # print(f"🧪 Rust match {case} side{i}: {rust_percent:.2f}%")
#             case_scores.append(round(rust_percent))

#         rust_scores_by_case[case] = case_scores

#     # Step 3: Return average rust match per case + individual side details
#     rust_summary = [round(np.mean(rust_scores_by_case[case])) for case in CASE_FOLDERS]
#     # print("📊 Final rust match summary:", rust_summary)

#     return jsonify({
#         "damage_per_image": rust_summary,
#         "details": {
#             case: {
#                 f"side{i+1}": rust_scores_by_case[case][i]
#                 for i in range(4)
#             } for case in CASE_FOLDERS
#         }
#     })

@app.route('/get_tasks', methods=['GET'])
def get_tasks():
    user_id = request.args.get('user_id')
    task_date = request.args.get('task_date')

    if not user_id or not task_date:
        return jsonify({"success": False, "message": "Missing parameters"}), 400

    # ✅ Use DictCursor to get row as dictionary
    connection = get_db_connection()
    cursor = connection.cursor(pymysql.cursors.DictCursor)

    cursor.execute("""
        SELECT * FROM assessment
        WHERE `User ID` = %s AND `Task Date` = %s
    """, (user_id, task_date))

    rows = cursor.fetchall()
    connection.close()

    return jsonify(rows)

@app.route('/get_remedies', methods=['GET'])
def get_remedies():
    user_id = request.args.get('user_id')
    task_date = request.args.get('task_date')

    if not user_id or not task_date:
        return jsonify({"success": False, "message": "Missing parameters"}), 400

    # ✅ Use DictCursor to get row as dictionary
    connection = get_db_connection()
    cursor = connection.cursor(pymysql.cursors.DictCursor)

    cursor.execute("""
        SELECT * FROM remedy
        WHERE `User ID` = %s AND `Task Date` = %s
    """, (user_id, task_date))

    rows = cursor.fetchall()
    connection.close()

    return jsonify(rows)

# ✅ Set your custom image storage path
UPLOAD_FOLDER = r"C:\Users\LENOVO\Desktop\AssessmentPictures"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route("/upload_single_image", methods=["POST"])
def upload_single_image():
    image = request.files.get("image")
    folder = request.form.get("folder")  # Assessment ID (used as subfolder)
    filename = request.form.get("filename")

    if not image or not folder or not filename:
        return "❌ Missing image or metadata", 400

    # ✅ Create subfolder based on Assessment ID
    folder_path = os.path.join(UPLOAD_FOLDER, folder)
    os.makedirs(folder_path, exist_ok=True)

    # ✅ Save the image
    image_path = os.path.join(folder_path, filename)
    image.save(image_path)

    return f"✅ Image saved to {folder}/{filename}", 200

# ✅ Set your custom Remedy image storage path
REMEDY_UPLOAD_FOLDER = r"C:\Users\LENOVO\Desktop\RemedyPictures"
os.makedirs(REMEDY_UPLOAD_FOLDER, exist_ok=True)

@app.route("/upload_remedysingle_image", methods=["POST"])
def upload_remedy_single_image():
    image = request.files.get("image")
    remedy_id = request.form.get("remedy_id")  # ✅ Still needed
    filename = request.form.get("filename")

    if not image or not remedy_id or not filename:
        return "❌ Missing image or metadata", 400

    # ✅ Create folder: RemedyPictures/<RemedyID>
    folder_path = os.path.join(REMEDY_UPLOAD_FOLDER, remedy_id)
    os.makedirs(folder_path, exist_ok=True)

    # ✅ Save the image inside the RemedyID folder
    image_path = os.path.join(folder_path, filename)
    image.save(image_path)

    return f"✅ Image saved to {remedy_id}/{filename}", 200


@app.route('/upload_tasks_to_pc', methods=['POST'])
def upload_tasks_to_pc():
    data = request.json
    uploaded_tasks = data.get("tasks", [])

    connection = get_db_connection()  # Your MySQL connection function
    cursor = connection.cursor(dictionary=True)  # ✅ mysql.connector-compatible

    updated = 0
    inserted = 0
    skipped = 0

    for row in uploaded_tasks:
        assessment_id = row[0]

        # ✅ Update Picture Location path for PC
        picture_location = rf"C:\Users\LENOVO\Desktop\AssessmentPictures\{assessment_id}"
        row[15] = picture_location  # Overwrite with correct PC path

        # Check if record exists
        cursor.execute("SELECT * FROM assessment WHERE `Assessment ID` = %s", (assessment_id,))
        existing = cursor.fetchone()

        if existing:
            # Compare each relevant field before deciding to update
            fields_to_compare = {
                "Task Date": row[5],
                "Allotted Date": row[6],
                "Allotted By": row[7],
                "Date Completed": row[8],
                "Assessment Status": row[9],
                "Assessment case": row[10],
                "Picture1 Name": row[11],
                "Picture2 Name": row[12],
                "Picture3 Name": row[13],
                "Picture4 Name": row[14],
                "Picture Location": row[15]
            }

            changes_required = False
            for key, new_value in fields_to_compare.items():
                if str(existing[key]) != str(new_value):
                    changes_required = True
                    break

            if changes_required:
                cursor.execute("""
                    UPDATE assessment SET
                        `Task Date` = %s,
                        `Allotted Date` = %s,
                        `Allotted By` = %s,
                        `Date Completed` = %s,
                        `Assessment Status` = %s,
                        `Assessment case` = %s,
                        `Picture1 Name` = %s,
                        `Picture2 Name` = %s,
                        `Picture3 Name` = %s,
                        `Picture4 Name` = %s,
                        `Picture Location` = %s
                    WHERE `Assessment ID` = %s
                """, (
                    row[5], row[6], row[7], row[8], row[9], row[10],
                    row[11], row[12], row[13], row[14], row[15],
                    assessment_id
                ))
                updated += 1
            else:
                skipped += 1  # No update needed
        else:
            # Insert new record
            cursor.execute("""
                INSERT INTO assessment (
                    `Assessment ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`, 
                    `Allotted Date`, `Allotted By`, `Date Completed`, `Assessment Status`, 
                    `Assessment case`, `Picture1 Name`, `Picture2 Name`, `Picture3 Name`, 
                    `Picture4 Name`, `Picture Location`
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, row)
            inserted += 1

    connection.commit()
    cursor.close()
    connection.close()

    return jsonify({
        "success": True,
        "message":f"Data Uploaded successfully --\n{inserted} inserted, {updated} updated, {skipped} skipped."
    })

@app.route('/upload_remedies_to_pc', methods=['POST'])
def upload_remedies_to_pc():
    data = request.json
    uploaded_remedies = data.get("remedies", [])

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    updated = 0
    inserted = 0
    skipped = 0

    for row in uploaded_remedies:
        remedy_id = row[0]
        area_id = row[1]

        # ✅ Update Picture Location path for PC
        picture_location = rf"C:\Users\LENOVO\Desktop\RemedyPictures\{remedy_id}"
        row[16] = picture_location  # Set Picture Location

        # Check if record exists
        cursor.execute("SELECT * FROM remedy WHERE `Remedy ID` = %s", (remedy_id,))
        existing = cursor.fetchone()

        if existing:
            fields_to_compare = {
                "Area ID": row[1],
                "User ID": row[2],
                "Table ID": row[3],
                "Pile No": row[4],
                "Task Date": row[5],
                "Allotted Date": row[6],
                "Allotted By": row[7],
                "Date Completed": row[8],
                "Assessed Case": row[9],
                "Remedy Status": row[10],
                "Remedy Text": row[11],
                "Picture1 Name": row[12],
                "Picture2 Name": row[13],
                "Picture3 Name": row[14],
                "Picture4 Name": row[15],
                "Picture Location": row[16]
            }

            changes_required = False
            for key, new_value in fields_to_compare.items():
                if str(existing.get(key)) != str(new_value):
                    changes_required = True
                    break

            if changes_required:
                cursor.execute("""
                    UPDATE remedy SET
                        `Area ID` = %s,
                        `User ID` = %s,
                        `Table ID` = %s,
                        `Pile No` = %s,
                        `Task Date` = %s,
                        `Allotted Date` = %s,
                        `Allotted By` = %s,
                        `Date Completed` = %s,
                        `Assessed Case` = %s,
                        `Remedy Status` = %s,
                        `Remedy Text` = %s,
                        `Picture1 Name` = %s,
                        `Picture2 Name` = %s,
                        `Picture3 Name` = %s,
                        `Picture4 Name` = %s,
                        `Picture Location` = %s
                    WHERE `Remedy ID` = %s
                """, (
                    row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8],
                    row[9], row[10], row[11], row[12], row[13], row[14], row[15], row[16],
                    remedy_id
                ))
                updated += 1
            else:
                skipped += 1
        else:
            cursor.execute("""
                INSERT INTO remedy (
                    `Remedy ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`,
                    `Allotted Date`, `Allotted By`, `Date Completed`, `Assessed Case`,
                    `Remedy Status`, `Remedy Text`, `Picture1 Name`, `Picture2 Name`,
                    `Picture3 Name`, `Picture4 Name`, `Picture Location`
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, row)
            inserted += 1

    connection.commit()
    cursor.close()
    connection.close()

    return jsonify({
        "success": True,
        "message": f"Remedies Uploaded Successfully --\n{inserted} inserted, {updated} updated, {skipped} skipped."
    })


# on your PC Flask server (app.py)
@app.route("/discover")
def discover():
    return jsonify({"status": "available", "ip": request.host.split(':')[0]})

@app.route('/sendtdpc')
def send_today_pc():
    return render_template('sendtdpc.html')

@app.route('/service-worker.js')
def serve_worker():
    return send_from_directory('static/js', 'service-worker.js')
    
@app.route('/')
def index():
    return render_template('login.html')

@app.route('/A001hotspot')
def A001hotspot():
    return render_template('A001hotspot.html')

@app.route('/A002hotspot')
def A002hotspot():
    return render_template('A002hotspot.html')

@app.route('/A003hotspot')
def A003hotspot():
    return render_template('A003hotspot.html')

@app.route('/A004hotspot')
def A004hotspot():
    return render_template('A004hotspot.html')

@app.route('/A005hotspot')
def A005hotspot():
    return render_template('A005hotspot.html')

@app.route('/A001remedyhotspot')
def A001_remedy_hotspot():
    return render_template('A001remedyhotspot.html')

@app.route('/A002remedyhotspot')
def A002_remedy_hotspot():
    return render_template('A002remedyhotspot.html')

@app.route('/A003remedyhotspot')
def A003_remedy_hotspot():
    return render_template('A003remedyhotspot.html')

@app.route('/A004remedyhotspot')
def A004_remedy_hotspot():
    return render_template('A004remedyhotspot.html')

@app.route('/A005remedyhotspot')
def A005_remedy_hotspot():
    return render_template('A005remedyhotspot.html')

# Route for the home page
@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'email' not in session:
        return redirect('/login')  # Redirect to login if not authenticated

    user_type = session.get('user_type', 'Normal User')
    username = session.get('username')  # Get the username from the session
    return render_template('dashboard.html', user_type=user_type, username=username)

@app.context_processor
def inject_user():
    full_name = session.get('username', '')  
    first_name = full_name.split()[0] if full_name else '' 
    return {
        'username': first_name,  
        'user_type': session.get('user_type', 'Normal User')
    }

# for create route 
@app.route('/site')
def site():
    return render_template('site.html')

# route to select the pile from the image map
@app.route('/image-map')
def image_map():
    return render_template('area1hotspot.html')

@app.route('/customer')
def customer():
    return render_template('customer.html')

@app.route('/userform')
def userform():
    return render_template('userform.html')

@app.route('/area')
def area():
    return render_template('area.html')

@app.route('/rows')
def rows():
    return render_template('rows.html')

@app.route('/tables')
def tables():
    return render_template('tables.html')

@app.route('/piles')
def piles():
    return render_template('pile.html')

@app.route('/assessment')
def assessment():
    return render_template('assessment.html')

@app.route('/bracingrust')
def bracingrust():
    return render_template('bracingrust.html')

@app.route('/remedy')
def remedy():
    return render_template('remedy.html')

@app.route('/inventory')
def inventory():
    return render_template('inventory.html')

@app.route('/invtrans')
def invtrans():
    return render_template('invtrans.html')

#@app.route('/quality')
#def quality():
    #return render_template('quality.html')

@app.route('/reports')
def reports():
    return render_template('assreports.html')

@app.route('/reporthtml')
def reporthtml():
    return render_template('reporthtml.html')
    
@app.route('/remedyreports')
def remedyreports():
    return render_template('remedyreports.html')

@app.route('/area-info')
def area_info():
    return render_template('area_info.html')

@app.route('/work-info')
def work_info():
    return render_template('workinfo.html')
@app.route('/remedywork-info')
def remedywork_info():
    return render_template('remedyworkinfo.html')

@app.route('/tableinfo')
def tableinfo():
    return render_template('tableinfo.html')
@app.route('/concreteimgrep')
def concreteimgrep():
    return render_template('concreteimgrep.html')

@app.route('/workallocation')
def workallocation():
    return render_template('workallocation.html')

@app.route('/profile', methods=['GET'])
def profile():
    if 'email' not in session:  
        return redirect('/login')  

    email = session['email']  # Retrieve the email from the session
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

   
    cursor.execute("SELECT * FROM users WHERE `Email` = %s", (email,))
    user = cursor.fetchone()

    if not user:  # Handle the case where no user is found
        return redirect('/login')  # Redirect to login if the user is not found

    # Pass the user data to the profile template
    return render_template('profile.html', user=user)


@app.route('/user_log')
def user_log():
    return render_template('user_log.html')

@app.route('/comments')
def comments():
    return render_template('comments.html')

@app.route('/area1remedyhotspot')
def area12hotspot():
    return render_template('area1remedyhotspot.html')

################################################################

#for update route

@app.route('/updateusers')
def update_users():
   
    return render_template('updateusers.html')

@app.route('/updatesite')
def update_site():
    
    return render_template('updatesite.html')

@app.route('/updatecustomer')
def update_customer():
   
    return render_template('updatecustomer.html')

@app.route('/updateinventory')
def update_inventory():
    
    return render_template('updateinventory.html')

@app.route('/updateinvtrans')
def update_invtrans():
   
    return render_template('updateinvtrans.html')

@app.route('/updatearea')
def update_area():
    return render_template('updatearea.html')

@app.route('/updatetable')
def update_table():
    return render_template('updatetable.html')

@app.route('/updatepile')
def update_pile():
    return render_template('updatepile.html')

@app.route('/updaterow')
def update_row():
    return render_template('updaterow.html')

@app.route('/updateassmnt')
def update_assmnt():
    return render_template('updateassmnt.html')

@app.route('/updateremedy')
def update_remedy():
    return render_template('updateremedy.html')

###########################################################
@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT * FROM users WHERE email = %s AND password = %s", (email, password))
        user = cursor.fetchone()

        if user:
            # Store necessary details in the session
            session['email'] = user['Email']  
            session['username'] = user['User Name']  
            session['user_type'] = user['User Type']

            return jsonify({"success": True, "message": "Login successful"})
        else:
            return jsonify({"success": False, "message": "Invalid credentials"})

    except Exception as e:
        return jsonify({"success": False, "message": "An error occurred", "error": str(e)})

    finally:
        cursor.close()
        connection.close()
 
@app.route('/get_assessment_data')
def get_assessment_data():
    table_id = request.args.get('table_id')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT 
            `Pile No` AS pile_no,
            `Assessment case` AS assessment_case,
            `Assessment Status` AS assessment_status
        FROM assessment
        WHERE `Table ID` = %s
        ORDER BY `Pile No`
    """
    cursor.execute(query, (table_id,))
    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(data)


@app.route('/update_assessment', methods=['POST'])
def update_assessment():
    data = request.json

    conn = get_db_connection()
    cursor = conn.cursor()

    query = """
        UPDATE assessment
        SET `Assessment case` = %s,
            `Assessment Status` = %s
        WHERE `Table ID` = %s
          AND `Pile No` = %s
    """

    cursor.execute(query, (
        data['assessment_case'],
        data['assessment_status'],
        data['table_id'],
        data['pile_no']
    ))

    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({
        "success": True,
        "message": f"Pile {data['pile_no']} updated successfully"
    })

# Route for user creation
@app.route('/create_user', methods=['POST'])
def create_user():
    data = request.get_json()
    name = data.get('name')
    email = data.get('email')
    password = data.get('password')

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500  # Handle connection failure

    cursor = connection.cursor()
    try:
        query = "INSERT INTO loginusers (name, email, password) VALUES (%s, %s, %s)"
        cursor.execute(query, (name, email, password))
        connection.commit()
        return jsonify({"success": True, "message": "User created successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error creating user: {e}"}), 500
    finally:
        cursor.close()  # Ensure cursor is closed
        connection.close()  # Ensure connection is closed

# Route for logging out
@app.route('/logout', methods=['POST'])
def logout():
    session.pop('username', None)  # Remove the username from the session
    return redirect(url_for('index'))

@app.route('/submit_siteform', methods=['POST'])
def submit_siteform():
    site_name = request.form.get('site_name')
    site_location = request.form.get('location')
    site_owner = request.form.get('site_owner_name')
    site_gps = request.form.get('site_gps')

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500  # Handle DB connection failure

    cursor = connection.cursor()
    try:
        # Fetch the current maximum Site ID
        cursor.execute("SELECT `Site ID` FROM `Site` ORDER BY `Site ID` DESC LIMIT 1")
        result = cursor.fetchone()
        
        next_number = int(result[0][1:]) + 1 if result else 1
        new_site_id = f"S{next_number:03d}"  

        query = """
        INSERT INTO `Site` (`Site ID`, `Cust ID`, `Site Name`, `Site Location`, `Site Owner Name`, `Site GPS`)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_site_id, "", site_name, site_location, site_owner, site_gps))
        connection.commit()

        return jsonify({"success": True, "message": "Site information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving site information: {e}"}), 500
    finally:
        cursor.close()  
        connection.close()  

@app.route('/submit_customerform', methods=['POST'])
def submit_customerform():
    name = request.form.get('name')
    address = request.form.get('address')
    contact_person = request.form.get('contact_person')
    website = request.form.get('website')
    phone_no = request.form.get('phone_no')
    country = request.form.get('country')

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()
    try:
        # Fetch the current maximum Cust ID
        cursor.execute("SELECT `Cust ID` FROM `Customer` ORDER BY `Cust ID` DESC LIMIT 1")
        result = cursor.fetchone()
        
        next_number = int(result[0][1:]) + 1 if result else 1
        new_cust_id = f"C{next_number:03d}"  

        query = """
        INSERT INTO Customer 
        (`Cust ID`, `Customer Name`, `Customer Address`, `Contact Person`, `Customer Website`, `Phone No`, `Country`)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_cust_id, name, address, contact_person, website, phone_no, country))
        connection.commit()

        return jsonify({"success": True, "message": "Customer information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving customer information: {e}"}), 500
    finally:
        cursor.close()  
        connection.close()  
@app.route('/submit_user_form', methods=['POST'])
def submit_userform():
    # Retrieve form data
    user_name = request.form.get('user_name')
    user_type = request.form.get('user_type')
    designation = request.form.get('designation')
    phone_no = request.form.get('phone_no')
    reports_to = request.form.get('reports_to')
    date_created = request.form.get('date_created')
    site_id = request.form.get('site_id')
    email = request.form.get('gmail_address')
    password = request.form.get('create_password')
    confirm_password = request.form.get('confirm_password')

    if password != confirm_password:
        return jsonify({"success": False, "message": "Passwords do not match!"})

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()
    try:
        # Check if the Site ID exists
        cursor.execute("SELECT `Site ID` FROM `site` WHERE `Site ID` = %s", (site_id,))
        if not cursor.fetchone():
            return jsonify({"success": False, "message": "Invalid Site ID selected."})

        # Check if Email already exists
        cursor.execute("SELECT `Email` FROM `users` WHERE `Email` = %s", (email,))
        if cursor.fetchone():
            return jsonify({"success": False, "message": "Email already exists. Please use a different email."})

        # ⬇️ NEW USER ID GENERATION LOGIC (U01, U02, U03...)
        cursor.execute("SELECT `User ID` FROM `users`")
        all_ids = cursor.fetchall()

        new_style_numbers = []
        for row in all_ids:
            uid = row[0]
            # Accept only U01, U02, ..., U99 (2-digit format)
            if uid.startswith("U") and len(uid) == 3 and uid[1:].isdigit():
                new_style_numbers.append(int(uid[1:]))

        if new_style_numbers:
            next_number = max(new_style_numbers) + 1
        else:
            next_number = 1

        new_user_id = f"U{next_number:02d}"
        # ⬆️ END OF NEW USER ID LOGIC

        query = """
        INSERT INTO `users` 
        (`User ID`, `Site ID`, `User Name`, `User Type`, `User Designation`, `User Phone number`, 
         `Reports To`, `Date Created`, `Date Removed`, `Email`, `Password`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(
            query,
            (new_user_id, site_id, user_name, user_type, designation,
             phone_no, reports_to, date_created, None, email, password)
        )
        connection.commit()

        return jsonify({"success": True, "message": "User information saved successfully"})
    
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving user information: {e}"}), 500
    
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_area_form', methods=['POST'])
def submit_area_form():
    # Retrieve form data
    location = request.form.get('location')
    gps = request.form.get('gps')
    if not location or not gps:
        return jsonify({"success": False, "message": "All fields are required."})

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        # Fetch the current maximum Area ID
        cursor.execute("SELECT `Area ID` FROM `areas` ORDER BY `Area ID` DESC LIMIT 1")
        result = cursor.fetchone()

        # Determine the next Area ID
        next_number = int(result[0][1:]) + 1 if result else 1
        new_area_id = f"A{next_number:03d}"

        query = """
        INSERT INTO `areas` (`Area ID`, `Location`, `GPS`)
        VALUES (%s, %s, %s)
        """
        cursor.execute(query, (new_area_id, location, gps)) 
        connection.commit()

        return jsonify({"success": True, "message": "Area information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving area information: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_user_log_form', methods=['POST'])
def submit_user_log_form():
    user_id = request.form.get('user_id')
    date_logged_in = request.form.get('date_logged_in')
    date_logged_out = request.form.get('date_logged_out')

    if not user_id or not date_logged_in:
        return jsonify({"success": False, "message": "User ID and Date Logged In are required."}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        query = """
        INSERT INTO `user_log` (`User ID`, `Date Logged in`, `Date Logged out`)
        VALUES (%s, %s, %s)
        """
        cursor.execute(query, (user_id, date_logged_in, date_logged_out))
        connection.commit()
        return jsonify({"success": True, "message": "User log information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving user log information: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_comment_form', methods=['POST'])
def submit_comment_form():
    required_fields = ["comment_type", "pile_id", "user_id", "date_posted", "comment_text"]
    data = {field: request.form.get(field) for field in required_fields}

    if not all(data.values()):
        return jsonify({"success": False, "message": "Required fields are missing!"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        query = """
        INSERT INTO `Comments` (`Comment Type`, `Related Comment ID`, `Pile ID`, `User ID`, `Usage ID`, 
                               `Date Posted`, `Comment Text`, `Comment Date`, `Commented By`, `Status`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (
            data["comment_type"], request.form.get("related_comment_id"), data["pile_id"], data["user_id"],
            request.form.get("usage_id"), data["date_posted"], data["comment_text"], request.form.get("comment_date"),
            request.form.get("commented_by"), request.form.get("status")
        ))
        connection.commit()
        return jsonify({"success": True, "message": "Comment information saved successfully"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving comment information: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

# @app.route('/submit_task_assignment', methods=['POST'])
# def submit_task_assignment():
#     area_id = request.form.get('area_id')
#     user_id = request.form.get('user_id')
#     table_id = request.form.get('selectedHotspots')  # Comma-separated Table IDs
#     task_date = request.form.get('task_date')
#     allotted_date = request.form.get('allotted_date')
#     allotted_by = request.form.get('allotted_by')
#     date_completed = request.form.get('date_completed')
#     confirm_insert = request.form.get('confirm_insert', 'no')  # Optional frontend input

#     if not table_id:
#         return jsonify({"success": False, "message": "No tables selected!"}), 400

#     table_id_list = [table.strip() for table in table_id.split(",") if table.strip()]

#     connection = get_db_connection()
#     if not connection:
#         return jsonify({"success": False, "message": "Database connection failed"}), 500

#     cursor = connection.cursor()

#     try:
#         existing_tables = []

#         for table in table_id_list:
#             cursor.execute("SELECT COUNT(*) FROM assessment WHERE `Table ID` = %s", (table,))
#             count = cursor.fetchone()[0]
#             if count > 0:
#                 existing_tables.append(table)

#         if existing_tables and confirm_insert.lower() == 'no':
#             return jsonify({
#                 "success": False,
#                 "message": f"Table ID(s) already present: {', '.join(existing_tables)}. ",
#                 "existing_tables": existing_tables
#             }), 409

#         for table in table_id_list:
#             if table in existing_tables and confirm_insert.lower() != 'yes':
#                 continue  # Skip if user did not confirm

#             for pile_no in range(1, 5):
#                 # Generate next Assessment ID
#                 cursor.execute("SELECT `Assessment ID` FROM `assessment` ORDER BY `Assessment ID` DESC LIMIT 1")
#                 result = cursor.fetchone()
#                 next_number = int(result[0][2:]) + 1 if result else 1
#                 new_assessment_id = f"AS{next_number:05d}"

#                 query = """
#                 INSERT INTO `assessment` 
#                     (`Assessment ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`, 
#                      `Allotted Date`, `Allotted By`, `Date Completed`, `Assessment Status`, `Assessment case`)
#                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
#                 """
#                 cursor.execute(query, (
#                     new_assessment_id, area_id, user_id, table, pile_no,
#                     task_date, allotted_date, allotted_by, date_completed,
#                     "In Progress", "Not Assessed"
#                 ))

#         connection.commit()
#         return jsonify({"success": True, "message": "Task assignment saved successfully"})

#     except mysql.connector.Error as e:
#         return jsonify({"success": False, "message": f"Error saving task assignment: {e}"}), 500
#     finally:
#         cursor.close()
#         connection.close()

@app.route('/submit_task_assignment', methods=['POST'])
def submit_task_assignment():
    area_id = request.form.get('area_id')
    user_id = request.form.get('user_id')
    table_id = request.form.get('selectedHotspots')  # Comma-separated Table IDs
    task_date = request.form.get('task_date')
    allotted_date = request.form.get('allotted_date')
    contractor = request.form.get('allotted_by')  # same variable, now maps to Contractor column
    date_completed = request.form.get('date_completed')
    confirm_insert = request.form.get('confirm_insert', 'no')

    if not table_id:
        return jsonify({"success": False, "message": "No tables selected!"}), 400

    table_id_list = [table.strip() for table in table_id.split(",") if table.strip()]

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        existing_tables = []

        for table in table_id_list:
            cursor.execute("SELECT COUNT(*) FROM assessment WHERE `Table ID` = %s", (table,))
            count = cursor.fetchone()[0]
            if count > 0:
                existing_tables.append(table)

        if existing_tables and confirm_insert.lower() == 'no':
            return jsonify({
                "success": False,
                "message": f"Table ID(s) already present: {', '.join(existing_tables)}.",
                "existing_tables": existing_tables
            }), 409

        for table in table_id_list:
            if table in existing_tables and confirm_insert.lower() != 'yes':
                continue

            for pile_no in range(1, 5):

                cursor.execute("SELECT `Assessment ID` FROM `assessment` ORDER BY `Assessment ID` DESC LIMIT 1")
                result = cursor.fetchone()
                next_number = int(result[0][2:]) + 1 if result else 1
                new_assessment_id = f"AS{next_number:05d}"

                # 🔥 Updated query – replaced `Allotted By` with `Contractor`
                query = """
                INSERT INTO `assessment` 
                    (`Assessment ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`, 
                     `Allotted Date`, `Contractor`, `Date Completed`, `Assessment Status`, `Assessment case`)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    new_assessment_id, area_id, user_id, table, pile_no,
                    task_date, allotted_date, contractor, date_completed,
                    "In Progress", "Not Assessed"
                ))

        connection.commit()
        return jsonify({"success": True, "message": "Task assignment saved successfully"})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving task assignment: {e}"}), 500

    finally:
        cursor.close()
        connection.close()
@app.route('/submit_rust_assignment', methods=['POST'])
def submit_rust_assignment():
    area_id = request.form.get('area_id')
    user_id = request.form.get('user_id')
    table_id = request.form.get('selectedHotspots')  # Comma-separated Table IDs
    task_date = request.form.get('task_date')
    allotted_date = request.form.get('allotted_date')
    contractor = request.form.get('allotted_by')  # same variable, now maps to Contractor column
    date_completed = request.form.get('date_completed')
    confirm_insert = request.form.get('confirm_insert', 'no')

    if not table_id:
        return jsonify({"success": False, "message": "No tables selected!"}), 400

    table_id_list = [table.strip() for table in table_id.split(",") if table.strip()]

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
        existing_tables = []

        for table in table_id_list:
            cursor.execute("SELECT COUNT(*) FROM bracingrust WHERE `Table ID` = %s", (table,))
            count = cursor.fetchone()[0]
            if count > 0:
                existing_tables.append(table)

        if existing_tables and confirm_insert.lower() == 'no':
            return jsonify({
                "success": False,
                "message": f"Table ID(s) already present: {', '.join(existing_tables)}.",
                "existing_tables": existing_tables
            }), 409

        for table in table_id_list:
            if table in existing_tables and confirm_insert.lower() != 'yes':
                continue

            for pile_no in range(1, 5):

                cursor.execute("SELECT `Assessment ID` FROM `bracingrust` ORDER BY `Assessment ID` DESC LIMIT 1")
                result = cursor.fetchone()
                next_number = int(result[0][2:]) + 1 if result else 1
                new_assessment_id = f"BR{next_number:05d}"

                # 🔥 Updated query – replaced `Allotted By` with `Contractor`
                query = """
                INSERT INTO `bracingrust` 
                    (`Assessment ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`, 
                     `Allotted Date`, `Contractor`, `Date Completed`, `Assessment Status`, `Assessment case`)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    new_assessment_id, area_id, user_id, table, pile_no,
                    task_date, allotted_date, contractor, date_completed,
                    "In Progress", "Not Assessed"
                ))

        connection.commit()
        return jsonify({"success": True, "message": "Task assignment saved successfully"})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving task assignment: {e}"}), 500

    finally:
        cursor.close()
        connection.close()

@app.route('/submit_remedy_form', methods=['POST'])
def submit_remedy_form():
    # Get form data
    area_id = request.form.get('area_id')
    user_id = request.form.get('user_id')
    table_ids = request.form.get('selectedHotspots')  # Comma-separated Table IDs
    task_date = request.form.get('task_date')
    assessed_case = request.form.get('assessed_case')
    allotted_date = request.form.get('allotted_date')
    allotted_by = request.form.get('allotted_by')
    date_completed = request.form.get('date_completed')
    remedy_status = request.form.get('remedy_status') or "In Progress"  # Default to "In Progress"
    remedy_text = request.form.get('remedy_text')

    if not table_ids:
        return jsonify({"success": False, "message": "No tables selected!"}), 400

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        table_id_list = [table.strip() for table in table_ids.split(",") if table.strip()]

        for table_id in table_id_list:
            for pile_no in range(1, 5):  # Create 4 rows per Table ID with Pile No 1-4
                # Fetch the latest Remedy ID
                cursor.execute("SELECT `Remedy ID` FROM `Remedy` ORDER BY `Remedy ID` DESC LIMIT 1")
                result = cursor.fetchone()

                next_number = int(result[0][2:]) + 1 if result else 1
                new_remedy_id = f"RM{next_number:05d}"

                # Insert each row with increasing Pile No
                query = """
                INSERT INTO `Remedy` (`Remedy ID`, `Area ID`, `User ID`, `Table ID`, `Pile No`, `Task Date`, `Assessed Case`, 
                                      `Allotted Date`, `Allotted By`, `Date Completed`, `Remedy Status`, `Remedy Text`)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (new_remedy_id, area_id, user_id, table_id, pile_no, task_date,  "Not Assessed",
                                       allotted_date, allotted_by, date_completed, "In Progress", remedy_text))

        connection.commit()
        return jsonify({"success": True, "message": "Remedy form submitted successfully"})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving remedy form: {e}"}), 500
    finally:
        cursor.close()
        connection.close()

from werkzeug.datastructures import FileStorage

@app.route('/update_assessment_pics', methods=['POST'])
def update_assessment_pics():
    try:
        # ✅ SAFELY read form data (works for all content types)
        user_id = request.form.get('user_id')
        task_date = request.form.get('task_date')
        table_ids = request.form.get('table_id')
        assessment_status = request.form.get('assessment_status')
        assessment_case = request.form.get('assessment_case')
        date_completed = request.form.get('date_completed')

        if not table_ids or "-" not in table_ids:
            return jsonify({"success": False, "message": "Invalid Table ID format"}), 400

        table_id, pile_part = table_ids.split("-", 1)
        pile_no = ''.join(filter(str.isdigit, pile_part))

        if not (user_id and task_date and table_id and pile_no):
            return jsonify({"success": False, "message": "Missing required fields"}), 400

        connection = get_db_connection()
        cursor = connection.cursor(buffered=True)

        # 🔹 Fetch Assessment ID and old pictures
        cursor.execute("""
            SELECT `Assessment ID`,
                   `Picture1 Name`, `Picture2 Name`,
                   `Picture3 Name`, `Picture4 Name`
            FROM assessment
            WHERE `User ID`=%s AND `Task Date`=%s
              AND `Table ID`=%s AND `Pile No`=%s
        """, (user_id, task_date, table_id, pile_no))

        row = cursor.fetchone()
        if not row:
            cursor.close()
            connection.close()
            return jsonify({"success": False, "message": "Assessment not found"}), 404

        assessment_id, pic1, pic2, pic3, pic4 = row
        cursor.close()

        # 📁 Folder
        assessment_folder = os.path.join(app.config["UPLOAD_FOLDER"], assessment_id)
        os.makedirs(assessment_folder, exist_ok=True)

        image_paths = [pic1, pic2, pic3, pic4]

        # 🔹 Handle 4 images safely
        for i in range(4):
            file = request.files.get(f'image{i+1}')

            if isinstance(file, FileStorage) and file.filename:
                filename = f"{assessment_id}_{table_id}_Pile{pile_no}_side{i+1}.jpg"
                file.save(os.path.join(assessment_folder, filename))
                image_paths[i] = filename

        # 🔹 Update DB
        cursor = connection.cursor()
        if date_completed:
            cursor.execute("""
                UPDATE assessment
                SET `Assessment Status`=%s,
                    `Assessment Case`=%s,
                    `Date Completed`=%s,
                    `Picture1 Name`=%s,
                    `Picture2 Name`=%s,
                    `Picture3 Name`=%s,
                    `Picture4 Name`=%s,
                    `Picture Location`=%s
                WHERE `Assessment ID`=%s
            """, (
                assessment_status, assessment_case, date_completed,
                image_paths[0], image_paths[1], image_paths[2], image_paths[3],
                assessment_folder, assessment_id
            ))
        else:
            cursor.execute("""
                UPDATE assessment
                SET `Assessment Status`=%s,
                    `Assessment Case`=%s,
                    `Picture1 Name`=%s,
                    `Picture2 Name`=%s,
                    `Picture3 Name`=%s,
                    `Picture4 Name`=%s,
                    `Picture Location`=%s
                WHERE `Assessment ID`=%s
            """, (
                assessment_status, assessment_case,
                image_paths[0], image_paths[1], image_paths[2], image_paths[3],
                assessment_folder, assessment_id
            ))

        connection.commit()
        cursor.close()
        connection.close()

        return jsonify({
            "success": True,
            "message": "Assessment updated successfully",
            "folder": assessment_folder
        })

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500

        
@app.route('/update_remedy_pics', methods=['POST'])
def update_remedy_pics():
    try:
        # Get form data
        user_id = request.form.get('user_id')
        task_date = request.form.get('task_date')
        table_ids = request.form.get('table_id')
        assessed_case = request.form.get('assessed_case')  
        remedy_status = request.form.get('remedy_status')
        date_completed = request.form.get('date_completed')

        if not table_ids or "-" not in table_ids:
            return jsonify({"success": False, "message": "Invalid Table ID format"}), 400

        table_id = table_ids.split("-")[0]
        pile_no = table_ids.split("-")[1][-1]

        if not (user_id and task_date and table_id and pile_no):
            return jsonify({"success": False, "message": "Missing required fields"}), 400

        # Get the corresponding Remedy ID
        connection = get_db_connection()
        cursor = connection.cursor(buffered=True)

        cursor.execute("""
            SELECT `Remedy ID`, `Picture1 Name`
            FROM remedy 
            WHERE `User ID` = %s AND `Task Date` = %s AND `Table ID` = %s AND `Pile No` = %s
        """, (user_id, task_date, table_id, pile_no))

        result = cursor.fetchone()
        cursor.close()

        if not result:
            connection.close()
            return jsonify({"success": False, "message": "Remedy not found"}), 404

        remedy_id = result[0]

        # Create a unique folder for the remedy
        remedy_folder = os.path.join(app.config["UPLOAD_IMG"], f"{remedy_id}")
        os.makedirs(remedy_folder, exist_ok=True)

        image_paths = []
        for i in range(1):
            image = request.files.get(f'image{i+1}')
            if image:
                image_filename = f"{remedy_id}_{table_id}_Pile{pile_no}_side{i+1}.jpg"
                image_path = os.path.join(remedy_folder, image_filename)
                image.save(image_path)
                image_paths.append(image_filename)
            else:
                image_paths.append(None)

        # Update database
        cursor = connection.cursor()
        query = """
            UPDATE remedy 
            SET `Remedy Status` = %s, `Assessed Case` = %s, `Date Completed` = %s,
                `Picture1 Name` = %s,
                `Picture Location` = %s
            WHERE `Remedy ID` = %s
        """

        cursor.execute(query, (
            remedy_status, assessed_case, date_completed,
            image_paths[0], 
            remedy_folder, remedy_id
        ))

        connection.commit()
        cursor.close()
        connection.close()

        return jsonify({"success": True, "message": "Remedy updated successfully", "folder": remedy_folder})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

   
@app.route('/get_submitted_hotspots', methods=['GET'])
def get_submitted_hotspots():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT `Table ID` FROM `assessment`")  
        rows = cursor.fetchall()

        # Flatten and split the stored values
        submitted_hotspots = []
        for row in rows:
            if row['Table ID']:
                submitted_hotspots.extend(row['Table ID'].split(","))  # Split on commas

        return jsonify({"submitted_hotspots": submitted_hotspots})  
    finally:
        cursor.close()
        connection.close()


@app.route('/get_final_submitted_hotspots', methods=['GET'])
def get_final_submitted_hotspots():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT `Table ID` FROM `assessment` WHERE `Assessment Status` = 'OE Approved'")  
        rows = cursor.fetchall()

        # Flatten and split the stored values
        final_submitted_hotspots = []
        for row in rows:
            if row['Table ID']:
                final_submitted_hotspots.extend(row['Table ID'].split(","))  # Split on commas

        return jsonify({"final_submitted_hotspots": final_submitted_hotspots})  
    finally:
        cursor.close()
        connection.close()

@app.route('/get_submitted_hotspots_remedy', methods=['GET'])
def get_submitted_hotspots_remedy():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT `Table ID` FROM `remedy`")  
        rows = cursor.fetchall()

        # Flatten and split the stored values
        submitted_hotspots = []
        for row in rows:
            if row['Table ID']:
                submitted_hotspots.extend(row['Table ID'].split(","))  # Split on commas

        return jsonify({"submitted_hotspots": submitted_hotspots})  
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_row_form', methods=['POST'])
def submit_row_form():
    # Retrieve form data
    row_name = request.form.get('row_name')
    area_id = request.form.get('area_id')
    location = request.form.get('location')
    gps = request.form.get('gps')

    # Validate required fields
    if not row_name or not area_id or not location or not gps:
        return jsonify({"success": False, "message": "All fields are required."})

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Fetch the current maximum Row ID
        cursor.execute("SELECT `Row ID` FROM `rows` ORDER BY `Row ID` DESC LIMIT 1")
        result = cursor.fetchone()

        # Determine the next Row ID
        if result and result[0]:
            last_row_id = result[0]
            next_number = int(last_row_id[1:]) + 1
        else:
            next_number = 1

        # Format the new Row ID as 'R001', 'R002', etc.
        new_row_id = f"R{next_number:03d}"

        # Insert data into the rows table with the generated Row ID
        query = """
        INSERT INTO `rows` (`Row ID`, `Row Name`, `Area ID`, `Location`, `GPS`)
        VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_row_id, row_name, area_id, location, gps))
        connection.commit()

        return jsonify({"success": True, "message": f"Row information saved successfully with ID {new_row_id}"})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error saving row information: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/submit_pile_form', methods=['POST'])
def submit_pile_form():
    # Get form data
    pile_ids = request.form.get('table_id')  # Example: "C63S24, C63S25"
    area_id = request.form.get('area_id')
    location_description = request.form.get('location_description')
    gps_location = request.form.get('gps_location')

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if not pile_ids:
            return jsonify({"success": False, "message": "No table selected"})

        pile_list = pile_ids.split(", ")  # Convert "C63S24, C63S25" → ["C63S24", "C63S25"]

        # Get the current highest Pile ID (convert to integer)
        cursor.execute("SELECT MAX(CAST(`Pile ID` AS UNSIGNED)) FROM `piles`")
        result = cursor.fetchone()
        next_pile_id = int(result[0]) + 1 if result and result[0] else 1  # Start from 1 if empty

        # Insert 4 piles per table
        for table_id in pile_list:
            for pile_no in range(1, 5):  # Create pile numbers 1–4
                query = """
                INSERT INTO `piles` (`Pile ID`, `Area ID`, `Table ID`, `Pile No`, `Location Description`, `GPS Location`)
                VALUES (%s, %s, %s, %s, %s, %s)
                """
                cursor.execute(query, (
                    str(next_pile_id),       # Pile ID as string (varchar)
                    area_id,
                    table_id,
                    pile_no,
                    location_description,
                    gps_location
                ))
                next_pile_id += 1  # Increment for next pile

        connection.commit()
        return jsonify({"success": True, "message": "Pile information saved successfully"})

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": f"Database error: {e}"})

    finally:
        cursor.close()
        connection.close()

@app.route('/submit_table_form', methods=['POST'])
def submit_table_form():
    # Get form data
    table_ids = request.form.get('table_id')  # This contains multiple table IDs as a string (e.g., "T1, T2, T3")
    area_id = request.form.get('area_id')
    location_description = request.form.get('location')
    gps_location = request.form.get('gps_location')

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if not table_ids:
            return jsonify({"success": False, "message": "No table IDs provided"})

        # Split the comma-separated table IDs into a list
        table_list = table_ids.split(", ")  # Convert the string of table_ids to a list

        # Insert each table as a new row
        for table_id in table_list:
            # Insert into the `tables` table
            query = """
            INSERT INTO `tables` (`Table ID`, `Area ID`, `Location`, `GPS`)
            VALUES (%s, %s, %s, %s)
            """
            cursor.execute(query, (table_id, area_id, location_description, gps_location))

        connection.commit()

        return jsonify({"success": True, "message": "Table information saved successfully"})

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": f"Error saving table information: {e}"})

    finally:
        cursor.close()
        connection.close()

@app.route('/submit_inventory_details', methods=['POST'])
def submit_inventory_details():
    # Get form data
    item_type = request.form.get('item_type')
    item_uom = request.form.get('item_uom')
    item_desc = request.form.get('item_desc')
    item_avl_qty = request.form.get('item_avl_qty')
    item_ror = request.form.get('item_ror') or None
    item_value = request.form.get('item_value') or None
    item_rate = request.form.get('item_rate') or None

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Generate the next Item ID
        cursor.execute("SELECT `Item ID` FROM `Inventory` ORDER BY `Item ID` DESC LIMIT 1")
        result = cursor.fetchone()
        
        if result and result[0]:
            last_item_id = result[0]
            next_number = int(last_item_id[1:]) + 1
        else:
            next_number = 1
        
        # Format the new Item ID as 'I001', 'I002', etc.
        new_item_id = f"I{next_number:04d}"

        # Insert data into Inventory table
        query = """
        INSERT INTO `Inventory` (`Item ID`, `Item Type`, `Item UOM`, `Item Desc`, 
                                 `Item Avl Qty`, `Item ROR`, `Item Value`, `Item Rate`)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_item_id, item_type, item_uom, item_desc, item_avl_qty, item_ror, item_value, item_rate))
        connection.commit()

     
        return jsonify({"success": True, "message": f"Item details saved successfully "})

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": f"Error saving item details: {e}"})

    finally:
        cursor.close()
        connection.close()

#invtrans
@app.route('/submit_item_transaction_form', methods=['POST'])
def submit_item_transaction_form():
    # Get form data
    item_type = request.form.get('item_type')
    trans_qty = request.form.get('trans_qty')
    trans_type = request.form.get('trans_type')
    trans_date = request.form.get('trans_date')
    user_id = request.form.get('user_id')
    usage = request.form.get('usage')

    # Establish DB connection
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Get the last Item ID
        cursor.execute("SELECT `Item ID` FROM `Invtrans` ORDER BY `Item ID` DESC LIMIT 1")
        result = cursor.fetchone()

        if result and result[0]:  # Ensure result exists and is not None
            last_item_id = result[0]

            if last_item_id.startswith("IT"):  # Check format
                try:
                    next_number = int(last_item_id[2:]) + 1  # Extract number part safely
                except ValueError:
                    next_number = 1  # Reset if format is incorrect
            else:
                next_number = 1  # Reset if format is not as expected
        else:
            next_number = 1  # If no previous records exist, start with IT001

        # Generate new Item ID
        new_item_id = f"IT{next_number:04d}"  # Format as IT001, IT002, etc.

        # Insert data into Item Transaction table
        query = """
        INSERT INTO `invtrans` (`Item ID`, `Item Type`, `Trans Qty`, `Trans Type`, 
                                `Trans Date`, `User ID`, `Usage`)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(query, (new_item_id, item_type, trans_qty, trans_type, trans_date, user_id, usage))
        connection.commit()

     
        return jsonify({"success": True, "message": "Item transaction details saved successfully"})

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "message": f"Error saving item transaction details: {e}"})

    finally:
        cursor.close()
        connection.close()

@app.route('/sync-db', methods=['POST'])
def sync_db():
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        tables = ['assessment', 'remedy']
        data = {}

        for table in tables:
            cursor.execute(f"SELECT * FROM {table} WHERE is_synced = FALSE")
            rows = cursor.fetchall()
            if rows:
                data[table] = rows

        if not data:
            return jsonify({"message": "No changes to sync."})

        res = requests.post("https://indisolar.tech/apply-changes", json=data)
        if res.status_code != 200:
            return jsonify({"message": "Server error applying changes."}), 500

        # Mark as synced
        for table, records in data.items():
            if records:
                ids = [str(r['id']) for r in records]
                cursor.execute(f"UPDATE {table} SET is_synced = TRUE WHERE id IN ({','.join(ids)})")
        db.commit()

        return jsonify({"message": "Changes synced successfully."})
    except Exception as e:
        return jsonify({"message": f"Error: {e}"}), 500


###########################################################################
@app.route('/generate_rustreport', methods=['GET'])
def generate_rustreport():
    user_id = request.args.get('user_id')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    completed_date = request.args.get('completed_date')  # ✅ New
    assessment_status = request.args.get('assessment_status')
    area_id = request.args.get('area')
    picture_location = request.args.get('picture_location')  # ✅ New

    if not completed_date and (not from_date or not to_date):
        return jsonify({"error": "Either Completed Date or From/To Date is required"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"error": "Database connection failed"}), 500

    cursor = connection.cursor(dictionary=True)

    try:
        query = """
            SELECT 
                a.`Assessment ID`, 
                u.`User Name`, 
                a.`Table ID`, 
                a.`Pile No`, 
                DATE_FORMAT(a.`Task Date`, '%d %b %Y') AS `Task Date`,
                a.`Assessment Status`, 
                a.`Assessment Case`,
                a.`Area ID`,
                a.`Date Completed`,
                a.`Picture Location`
            FROM bracingrust a
            JOIN users u ON a.`User ID` = u.`User ID`
            WHERE 1 = 1
        """
        params = []

        # ✅ Add date filter: Either Task Date range OR Completed Date
        if completed_date:
            query += " AND DATE(a.`Date Completed`) = %s"
            params.append(completed_date)
        else:
            query += " AND DATE(a.`Task Date`) BETWEEN %s AND %s"
            params.extend([from_date, to_date])

        if user_id and user_id.lower() != "all":
            query += " AND a.`User ID` = %s"
            params.append(user_id)

        if assessment_status and assessment_status.lower() != "all status":
            query += " AND a.`Assessment Status` = %s"
            params.append(assessment_status)

        if area_id and area_id.lower() != "all":
            query += " AND a.`Area ID` = %s"
            params.append(area_id)

        if picture_location:
            if picture_location == "not_null":
                query += " AND a.`Picture Location` IS NOT NULL AND a.`Picture Location` != ''"
            elif picture_location == "null":
                query += " AND (a.`Picture Location` IS NULL OR a.`Picture Location` = '')"

        cursor.execute(query, params)
        result = cursor.fetchall()

        return jsonify(result)

    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e}"}), 500

    finally:
        cursor.close()
        connection.close()

@app.route('/generate_report', methods=['GET'])
def generate_report():
    user_id = request.args.get('user_id')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    completed_date = request.args.get('completed_date')  # ✅ New
    assessment_status = request.args.get('assessment_status')
    area_id = request.args.get('area')
    picture_location = request.args.get('picture_location')  # ✅ New

    if not completed_date and (not from_date or not to_date):
        return jsonify({"error": "Either Completed Date or From/To Date is required"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"error": "Database connection failed"}), 500

    cursor = connection.cursor(dictionary=True)

    try:
        query = """
            SELECT 
                a.`Assessment ID`, 
                u.`User Name`, 
                a.`Table ID`, 
                a.`Pile No`, 
                DATE_FORMAT(a.`Task Date`, '%d %b %Y') AS `Task Date`,
                a.`Assessment Status`, 
                a.`Assessment Case`,
                a.`Area ID`,
                a.`Date Completed`,
                a.`Picture Location`
            FROM assessment a
            JOIN users u ON a.`User ID` = u.`User ID`
            WHERE 1 = 1
        """
        params = []

        # ✅ Add date filter: Either Task Date range OR Completed Date
        if completed_date:
            query += " AND DATE(a.`Date Completed`) = %s"
            params.append(completed_date)
        else:
            query += " AND DATE(a.`Task Date`) BETWEEN %s AND %s"
            params.extend([from_date, to_date])

        if user_id and user_id.lower() != "all":
            query += " AND a.`User ID` = %s"
            params.append(user_id)

        if assessment_status and assessment_status.lower() != "all status":
            query += " AND a.`Assessment Status` = %s"
            params.append(assessment_status)

        if area_id and area_id.lower() != "all":
            query += " AND a.`Area ID` = %s"
            params.append(area_id)

        if picture_location:
            if picture_location == "not_null":
                query += " AND a.`Picture Location` IS NOT NULL AND a.`Picture Location` != ''"
            elif picture_location == "null":
                query += " AND (a.`Picture Location` IS NULL OR a.`Picture Location` = '')"

        cursor.execute(query, params)
        result = cursor.fetchall()

        return jsonify(result)

    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e}"}), 500

    finally:
        cursor.close()
        connection.close()

@app.route('/update_assessment_case', methods=['POST'])
def update_assessment_case():
    data = request.get_json()

    conn = get_db_connection()
    cursor = conn.cursor()

    for item in data:
        table_id = item['table_id']
        pile_no = item['pile_no']
        new_case = item['assessment_case']
        new_status = item.get('assessment_status', None)

        if new_status:
            cursor.execute("""
                UPDATE assessment
                SET `Assessment case` = %s, `Assessment Status` = %s
                WHERE `Table ID` = %s AND `Pile No` = %s
            """, (new_case, new_status, table_id, pile_no))
        else:
            cursor.execute("""
                UPDATE assessment
                SET `Assessment case` = %s
                WHERE `Table ID` = %s AND `Pile No` = %s
            """, (new_case, table_id, pile_no))

    conn.commit()
    conn.close()

    return jsonify({"message": "Assessment Case(s) and Status updated successfully."})

@app.route('/get-area-info', methods=['POST'])
def get_area_info():
    data = request.get_json()
    selected_area = data.get('area')
    selected_table = data.get('table')
    start_date = data.get('start_date')  # optional
    end_date = data.get('end_date')      # optional
    user_id = data.get('user_id')        # ⬅️ NEW (optional)

    area_map = {
        "Area1": "A001",
        "Area2": "A002",
        "Area3": "A003",
        "Area4": "A004",
        "Area5": "A005"
    }

    db_area_id = area_map.get(selected_area)
    if not db_area_id:
        return jsonify({"success": False, "message": "Invalid area"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({"success": False, "message": "Database connection failed"}), 500

    cursor = connection.cursor()

    try:
      
        where_clause = "`Area ID` = %s"
        params = [db_area_id]


        if user_id:
            where_clause += " AND `User ID` = %s"
            params.append(user_id)

      
        if start_date and end_date:
            where_clause += " AND `Task Date` BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        if selected_table == "assessment":
            query = f"""
                SELECT
                    SUM(CASE WHEN `Assessment Status` = 'In Progress' THEN 1 ELSE 0 END) AS in_progress,
                    SUM(CASE WHEN `Assessment Status` = 'Completed' THEN 1 ELSE 0 END) AS completed,
                    SUM(CASE WHEN `Assessment Status` = 'PM Approved' THEN 1 ELSE 0 END) AS pm_approved,
                    SUM(CASE WHEN `Assessment Status` = 'OE Approved' THEN 1 ELSE 0 END) AS oe_approved,
                    SUM(CASE WHEN `Assessment Case` = 'Case1' THEN 1 ELSE 0 END) AS case1,
                    SUM(CASE WHEN `Assessment Case` = 'Case2' THEN 1 ELSE 0 END) AS case2,
                    SUM(CASE WHEN `Assessment Case` = 'Case3' THEN 1 ELSE 0 END) AS case3,
                    SUM(CASE WHEN `Assessment Case` = 'Case4' THEN 1 ELSE 0 END) AS case4
                FROM assessment
                WHERE {where_clause}
            """

        elif selected_table == "remedy":
            query = f"""
                SELECT
                    SUM(CASE WHEN `Remedy Status` = 'In Progress' THEN 1 ELSE 0 END) AS in_progress,
                    SUM(CASE WHEN `Remedy Status` = 'Completed' THEN 1 ELSE 0 END) AS completed,
                    SUM(CASE WHEN `Remedy Status` = 'PM Approved' THEN 1 ELSE 0 END) AS pm_approved,
                    SUM(CASE WHEN `Remedy Status` = 'OE Approved' THEN 1 ELSE 0 END) AS oe_approved,
                    SUM(CASE WHEN `Assessed Case` = 'Case1' THEN 1 ELSE 0 END) AS case1,
                    SUM(CASE WHEN `Assessed Case` = 'Case2' THEN 1 ELSE 0 END) AS case2,
                    SUM(CASE WHEN `Assessed Case` = 'Case3' THEN 1 ELSE 0 END) AS case3,
                    SUM(CASE WHEN `Assessed Case` = 'Case4' THEN 1 ELSE 0 END) AS case4
                FROM remedy
                WHERE {where_clause}
            """

        else:
            return jsonify({"success": False, "message": "Invalid table selected"}), 400

        cursor.execute(query, tuple(params))
        result = cursor.fetchone()

        response_data = {
            "success": True,
            "in_progress": result[0] or 0,
            "completed": result[1] or 0,
            "pm_approved": result[2] or 0,
            "oe_approved": result[3] or 0,
            "case1": result[4] or 0,
            "case2": result[5] or 0,
            "case3": result[6] or 0,
            "case4": result[7] or 0
        }

        return jsonify(response_data)

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        cursor.close()
        connection.close()

@app.route('/get_remedy_texts', methods=['GET'])
def get_remedy_texts():
    if "user_type" not in session:
        return jsonify({"success": False, "message": "User not authenticated"}), 401

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if session["user_type"] == "Admin":
            cursor.execute("""
                SELECT DISTINCT `Remedy Text`
                FROM remedy
                WHERE `Remedy Text` IS NOT NULL
                  AND `Remedy Text` != ''
            """)
        else:
            cursor.execute("""
                SELECT DISTINCT `Remedy Text`
                FROM remedy
                WHERE `Remedy Text` IS NOT NULL
                  AND `Remedy Text` != ''
                  AND `User ID` = %s
            """, (session["user_id"],))

        remedy_texts = [
            {"id": row[0], "username": row[0]}
            for row in cursor.fetchall()
        ]

        return jsonify({"success": True, "users": remedy_texts})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching Remedy Text: {e}"})

    finally:
        cursor.close()
        connection.close()
@app.route('/generate_remedy_report', methods=['GET'])
def generate_remedy_report():
    remedy_text = request.args.get('remedy_text')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    completed_date = request.args.get('completed_date')
    remedy_status = request.args.get('remedy_status')
    remedy_case = request.args.get('remedy_case')
    area_id = request.args.get('area')
    picture_location = request.args.get('picture_location')

    # Validation: Either date range OR completed date
    if (from_date and to_date and completed_date) or ((not from_date or not to_date) and not completed_date):
        return jsonify({
            "error": "Please either provide From + To Dates or a Completed Date — not both."
        }), 400

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        query = """
            SELECT 
                r.`Remedy ID`,
                r.`Area ID`,
                COALESCE(a.`Assessment Case`, 'Not Assessed') AS `Assessed Case`,
                COALESCE(a.`Assessment Status`, 'Not Assessed') AS `Remedy Status`,
                r.`Pile No`,
                r.`Table ID`,
                DATE_FORMAT(r.`Task Date`, '%d %b %Y') AS `Task Date`,
                r.`Remedy Text`
            FROM remedy r
            LEFT JOIN assessment a 
                ON r.`Area ID` = a.`Area ID`
               AND r.`Table ID` = a.`Table ID`
               AND r.`Pile No` = a.`Pile No`
            WHERE 1=1
        """

        params = []

        # Date filter
        if completed_date:
            query += " AND DATE(r.`Date Completed`) = %s"
            params.append(completed_date)
        else:
            query += " AND DATE(r.`Task Date`) BETWEEN %s AND %s"
            params.extend([from_date, to_date])

        # Remedy Text filter
        if remedy_text and remedy_text.lower() != "all":
            query += " AND r.`Remedy Text` = %s"
            params.append(remedy_text)

        # Remedy Status filter (NULL-safe)
        if remedy_status and remedy_status.lower() != "all status":
            query += """
                AND (
                    a.`Assessment Status` = %s
                    OR a.`Assessment Status` IS NULL
                )
            """
            params.append(remedy_status)

        # Remedy Case filter (NULL-safe)
        if remedy_case and remedy_case.lower() != "all cases":
            query += """
                AND (
                    a.`Assessment Case` = %s
                    OR a.`Assessment Case` IS NULL
                )
            """
            params.append(remedy_case)

        # Area filter
        if area_id and area_id.lower() != "all":
            query += " AND r.`Area ID` = %s"
            params.append(area_id)

        # Picture Location filter
        if picture_location == "not_null":
            query += " AND r.`Picture Location` IS NOT NULL AND r.`Picture Location` != ''"
        elif picture_location == "null":
            query += " AND (r.`Picture Location` IS NULL OR r.`Picture Location` = '')"

        # Final execution
        query += " ORDER BY r.`Remedy ID` ASC"

        cursor.execute(query, params)
        result = cursor.fetchall()

        return jsonify(result)

    finally:
        cursor.close()
        connection.close()


@app.route('/search_by_remedydate', methods=['GET'])
def search_by_remedydate():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    
    user_id = request.args.get('user_id')
    task_date = request.args.get('date')

    if not user_id or not task_date:
        return jsonify({"error": "User ID and Date parameters are required"}), 400

    try:
        # Fetch Table ID, Pile No, Remedy Status, and Assessed Case
        cursor.execute("""
            SELECT `Table ID`, `Pile No`, `Remedy Status`, `Assessed Case`
            FROM remedy
            WHERE `User ID` = %s 
            AND DATE(`Task Date`) = %s  
           
            ORDER BY `Table ID`, `Pile No`
        """, (user_id, task_date))

        result = cursor.fetchall()

        formatted_data = []
        for row in result:
            table_id = row["Table ID"]
            pile_no = row["Pile No"]

            formatted_data.append({
                "Assessed Case": row["Assessed Case"],
                "Remedy Status": row["Remedy Status"],
                "Pile No": pile_no,
                "Table ID": f"{table_id}-Pile{pile_no}"  # Format Table ID with Pile No
            })

        return jsonify({"data": formatted_data}) if formatted_data else jsonify({"data": []})

    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e}"}), 500

    finally:
        cursor.close()
        connection.close()
        
# @app.route('/generate_ass_slno_txt', methods=['POST'])
# def generate_ass_slno_txt():
#     data = request.get_json()
#     slno_type = request.args.get('type', 'assessment')  # 'normal' or 'case'
    
#     # File paths
#     txt_file_path = 'Ass_slno.txt'
#     excel_file_path = os.path.expanduser("~/Desktop/AssessmentSerial_No.xlsx")
#     last_slno_path = 'last_ass_slno.txt'

#     # Load or initialize last_slno.json
#     if os.path.exists(last_slno_path):
#         with open(last_slno_path, 'r') as f:
#             last_slno_data = json.load(f)
#     else:
#         last_slno_data = {"remedy": 0, "assessment": 0}

#     last_slno = last_slno_data.get(slno_type, 0)

#     # Step 1: Write to text file
#     with open(txt_file_path, 'a', encoding='utf-8') as f:
#         for item in data:
#             f.write(f"{last_slno}\t{item.get('assessment_id', '')}\t{item.get('area_id', '')}\t{item.get('table_id', '')}\t{item.get('pile_no', '')}\t{item.get('task_date', '')}\n")
#             last_slno += 1

#     # Step 2: Write to Excel
#     if os.path.exists(excel_file_path):
#         wb = load_workbook(excel_file_path)
#         ws = wb.active
#     else:
#         wb = Workbook()
#         ws = wb.active
#         ws.append(["Sl.No", "Assessment ID", "Area ID", "Table ID", "Pile No", "Task Date"])



#     for item in data:
#         ws.append([
#             last_slno - len(data) + data.index(item) + 1,
#             item.get('assessment_id', ''),
#             item.get('area_id', ''),
#             item.get('table_id', ''),
#             item.get('pile_no', ''),
#             item.get('task_date', '')
#         ])

#     try:
#         wb.save(excel_file_path)
#     except PermissionError:
#         return jsonify({"error": "Close AssessmentSerial_No.xlsx before generating."}), 500

#     # Step 3: Save updated SLNO to JSON
#     last_slno_data[slno_type] = last_slno
#     with open(last_slno_path, 'w') as f:
#         json.dump(last_slno_data, f)

#     return jsonify({"message": f"SLNO saved. Last SLNO for '{slno_type}': {last_slno}"})


@app.route('/generate_ass_slno_txt', methods=['POST'])
def generate_ass_slno_txt():
    data = request.get_json()
    slno_type = request.args.get('type', 'assessment')  # 'assessment', 'remedy', or 'case'
    
    # File paths
    txt_file_path = 'Ass_slno.txt'
    excel_file_path = os.path.expanduser("~/Desktop/AssessmentSerial_No.xlsx")
    last_slno_path = 'last_ass_slno.txt'

    # Load or initialize SLNO storage
    if os.path.exists(last_slno_path):
        with open(last_slno_path, 'r') as f:
            last_slno_data = json.load(f)
    else:
        last_slno_data = {"remedy": 0, "assessment": 0, "case": 0}

    last_slno = last_slno_data.get(slno_type, 0)

    # Step 1: Write to text file
    with open(txt_file_path, 'a', encoding='utf-8') as f:
        for item in data:
            f.write(
                f"{last_slno}\t{item.get('assessment_id', '')}\t{item.get('area_id', '')}\t"
                f"{item.get('table_id', '')}\t{item.get('pile_no', '')}\t{item.get('assessment_case', '')}\t"
                f"{item.get('task_date', '')}\n"
            )
            last_slno += 1

    # Step 2: Write to Excel
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Sl.No", "Assessment ID", "Area ID", "Table ID", "Pile No", "Case", "Task Date"])

    for idx, item in enumerate(data):
        ws.append([
            last_slno - len(data) + idx + 1,
            item.get('assessment_id', ''),
            item.get('area_id', ''),
            item.get('table_id', ''),
            item.get('pile_no', ''),
            item.get('assessment_case', ''),  # New Case column
            item.get('task_date', '')
        ])

    try:
        wb.save(excel_file_path)
    except PermissionError:
        return jsonify({"error": "Close AssessmentSerial_No.xlsx before generating."}), 500

    # Step 3: Save updated SLNO to JSON
    last_slno_data[slno_type] = last_slno
    with open(last_slno_path, 'w') as f:
        json.dump(last_slno_data, f)

    return jsonify({"message": f"SLNO saved. Last SLNO for '{slno_type}': {last_slno}"})

@app.route('/generate_slno_txt', methods=['POST'])
def generate_slno_txt():
    data = request.get_json()
    slno_type = request.args.get('type', 'remedy')  # 'normal' or 'case'
    
    # File paths
    txt_file_path = 'slno.txt'
    excel_file_path = os.path.expanduser("~/Desktop/RemedySerial_No.xlsx")
    last_slno_path = 'last_ass_slno.txt'

    # Load or initialize last_slno.json
    if os.path.exists(last_slno_path):
        with open(last_slno_path, 'r') as f:
            last_slno_data = json.load(f)
    else:
        last_slno_data = {"remedy": 0, "assessment": 0}

    last_slno = last_slno_data.get(slno_type, 0)

    # Step 1: Write to text file
    with open(txt_file_path, 'a') as f:
        for item in data:
            last_slno += 1
            f.write(f"{last_slno}\t{item['remedy_id']}\t{item['table_id']}\t{item['pile_no']}\n")

    # Step 2: Write to Excel
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Sl.No", "Remedy ID", "Table ID", "Pile No"])

    for item in data:
        ws.append([
            last_slno - len(data) + data.index(item) + 1,
            item['remedy_id'],
            item['table_id'],
            item['pile_no']
        ])

    try:
        wb.save(excel_file_path)
    except PermissionError:
        return jsonify({"error": "Close remSerial_no.xlsx before generating."}), 500

    # Step 3: Save updated SLNO to JSON
    last_slno_data[slno_type] = last_slno
    with open(last_slno_path, 'w') as f:
        json.dump(last_slno_data, f)

    return jsonify({"message": f"SLNO saved. Last SLNO for '{slno_type}': {last_slno}"})

# Helper to read last slno based on a specific type (case or normal)
def get_last_slno(slno_type="case"):
    slno_file = "last_slno.txt"
    try:
        with open(slno_file, "r") as f:
            data = json.load(f)
    except Exception as e:
        data = {}
        print(f"Error reading slno file: {e}")

    return data.get(slno_type, 0)  # Return the last number for the specified type, default to 0 if not found

# Helper to update last slno for a specific type (case or normal)
def update_last_slno(slno_type, new_slno):
    slno_file = "last_slno.txt"
    try:
        # Read existing data
        if os.path.exists(slno_file):
            with open(slno_file, "r") as f:
                data = json.load(f)
        else:
            data = {}

        # Update the specified slno_type key
        data[slno_type] = new_slno

        # Write updated data back to file
        with open(slno_file, "w") as f:
            json.dump(data, f)

    except Exception as e:
        print(f"Error updating slno file: {e}")

# Route to get the last Sl. No. based on a specific type (case or normal)
@app.route('/get_last_slno', methods=['GET'])
def get_slno():
    slno_type = request.args.get("type", "case")  # Default to "case" if no type is provided
    return jsonify({"last_slno": get_last_slno(slno_type)})
# Route to save the Remedy Pics PDF
@app.route('/save_remedypics_pdf', methods=['POST'])
def save_remedypics_pdf():
    pdf_file = request.files['pdf']
    new_last_slno = int(request.form.get("new_last_slno", "0"))

    # Get Desktop Path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Create "Remedy Reports/Pics Reports" Folder if it doesn't exist
    reports_folder = os.path.join(desktop_path, "Remedy Reports")
    os.makedirs(reports_folder, exist_ok=True)

    # Prepare file name
    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)
    pdf_path = os.path.join(reports_folder, original_filename)

    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    pdf_file.save(pdf_path)

    # 🔁 Update only the "remedypics" key in last_slno.txt
    update_last_slno("remedypics", new_last_slno)

    return jsonify({"message": f"Remedy Pics PDF saved as {os.path.basename(pdf_path)}!"})


# Route to save the Remedy Case PDF
@app.route('/save_remedycase_pdf', methods=['POST'])
def save_remedycase_pdf():
    pdf_file = request.files['pdf']
    new_last_slno = int(request.form.get("new_last_slno", "0"))

    # Get Desktop Path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Create "Remedy Reports/Case Reports" Folder if it doesn't exist
    reports_folder = os.path.join(desktop_path, "Remedy Reports", "Case Reports")
    os.makedirs(reports_folder, exist_ok=True)

    # Prepare file name
    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)
    pdf_path = os.path.join(reports_folder, original_filename)

    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    pdf_file.save(pdf_path)

    # 🔁 Update only the "remedycase" key in last_slno.txt
    update_last_slno("remedycase", new_last_slno)

    return jsonify({"message": f"Remedy Case PDF saved as {os.path.basename(pdf_path)}!"})

@app.route('/saveremedy_pdf', methods=['POST'])
def saveremedy_pdf():
    pdf_file = request.files['pdf']
    new_last_slno = int(request.form.get("new_last_slno", "0"))

    # Desktop path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    reports_folder = os.path.join(desktop_path, "Remedy Reports")
    os.makedirs(reports_folder, exist_ok=True)

    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)
    pdf_path = os.path.join(reports_folder, original_filename)

    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    pdf_file.save(pdf_path)

    update_last_slno("remedy", new_last_slno)

    return jsonify({"message": f"Remedy PDF saved successfully as {os.path.basename(pdf_path)}!"})

# Route to save the normal PDF
@app.route('/save_pdf', methods=['POST'])
def save_pdf():
    pdf_file = request.files['pdf']
    new_last_slno = int(request.form.get("new_last_slno", "0"))

    # Get Desktop Path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Create "Assessment Reports" Folder if it doesn't exist
    reports_folder = os.path.join(desktop_path, "Assessment Reports")
    os.makedirs(reports_folder, exist_ok=True)

    # Extract base name and extension
    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)

    # Prepare file path and check for duplicates
    pdf_path = os.path.join(reports_folder, original_filename)
    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    # Save the file with a unique name
    pdf_file.save(pdf_path)

    # 🔁 Update only the "normal" key in last_slno.txt
    slno_type = "normal"
    update_last_slno(slno_type, new_last_slno)

    return jsonify({"message": f"PDF saved successfully as {os.path.basename(pdf_path)}!"})

@app.route('/save_rustpdf', methods=['POST'])
def save_rustpdf():
    pdf_file = request.files['pdf']
    new_last_slno = int(request.form.get("new_last_slno", "0"))

    # Get Desktop Path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Create "Assessment Reports" Folder if it doesn't exist
    reports_folder = os.path.join(desktop_path, "Assessment Reports")
    os.makedirs(reports_folder, exist_ok=True)

    # Extract base name and extension
    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)

    # Prepare file path and check for duplicates
    pdf_path = os.path.join(reports_folder, original_filename)
    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    # Save the file with a unique name
    pdf_file.save(pdf_path)

    # 🔁 Update only the "normal" key in last_slno.txt
    slno_type = "bracingrust"
    update_last_slno(slno_type, new_last_slno)

    return jsonify({"message": f"PDF saved successfully as {os.path.basename(pdf_path)}!"})

# Route to save the case PDF
@app.route('/savecase_pdf', methods=['POST'])
def savecase_pdf():
    pdf_file = request.files['pdf']
    new_last_slno = int(request.form.get("new_last_slno", "0"))

    # Get Desktop Path
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Create "Assessment Reports" Folder if it doesn't exist
    reports_folder = os.path.join(desktop_path, "Assessment Reports", "Case Reports")
    os.makedirs(reports_folder, exist_ok=True)

    # Extract base name and extension
    original_filename = pdf_file.filename
    base_name, extension = os.path.splitext(original_filename)

    # Prepare file path and check for duplicates
    pdf_path = os.path.join(reports_folder, original_filename)
    count = 1
    while os.path.exists(pdf_path):
        pdf_path = os.path.join(reports_folder, f"{base_name}_{count}{extension}")
        count += 1

    # Save the file with a unique name
    pdf_file.save(pdf_path)

    # 🔁 Update only the "case" key in last_slno.txt
    slno_type = "case"
    update_last_slno(slno_type, new_last_slno)

    return jsonify({"message": f"PDF saved successfully as {os.path.basename(pdf_path)}!"})


@app.route('/get_user_ids', methods=['GET'])
def get_user_ids():
    if "user_type" not in session:
        return jsonify({"success": False, "message": "User not authenticated"}), 401

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if session["user_type"] == "Admin":
            # Admin can see all users
            cursor.execute("SELECT `User ID`, `User Name` FROM `users`")
        else:
            # Normal user can see only their own details
            cursor.execute("SELECT `User ID`, `User Name` FROM `users` WHERE `Email` = %s", (session["email"],))

        users = [{"id": row[0], "username": row[1]} for row in cursor.fetchall()]
        
        return jsonify({"success": True, "users": users})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching User IDs: {e}"})

    finally:
        cursor.close()
        connection.close()
@app.route('/get_contractors', methods=['GET'])
def get_contractors():
    if "user_type" not in session:
        return jsonify({"success": False, "message": "User not authenticated"}), 401

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        if session["user_type"] == "Admin":
            cursor.execute("""
                SELECT DISTINCT u.`User ID`, u.`User Name`
                FROM assessment a
                JOIN users u ON a.Contractor = u.`User ID`
                WHERE a.Contractor IS NOT NULL
                  AND a.Contractor != ''
            """)
        else:
            cursor.execute("""
                SELECT DISTINCT u.`User ID`, u.`User Name`
                FROM assessment a
                JOIN users u ON a.Contractor = u.`User ID`
                WHERE a.Contractor IS NOT NULL
                  AND a.Contractor != ''
                  AND a.`User ID` = %s
            """, (session["user_id"],))

        contractors = [
            {"id": row[0], "username": row[1]}
            for row in cursor.fetchall()
        ]

        return jsonify({"success": True, "users": contractors})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching contractors: {e}"})

    finally:
        cursor.close()
        connection.close()


# @app.route('/get_user_ids', methods=['GET'])
# def get_user_ids():
#     if "user_type" not in session:
#         return jsonify({"success": False, "message": "User not authenticated"}), 401

#     connection = get_db_connection()
#     cursor = connection.cursor()

#     try:
#         # ONLY return U01 and U02
#         cursor.execute("""
#             SELECT `User ID`, `User Name` 
#             FROM `users`
#             WHERE `User ID` IN ('U01','U02')
#         """)

#         users = [{"id": row[0], "username": row[1]} for row in cursor.fetchall()]
        
#         return jsonify({"success": True, "users": users})

#     except mysql.connector.Error as e:
#         return jsonify({"success": False, "message": f"Error fetching User IDs: {e}"})

#     finally:
#         cursor.close()
#         connection.close()


@app.route('/get_site_ids', methods=['GET'])
def get_site_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        cursor.execute("SELECT `Site ID`, `Site Name` FROM `site`")
        sites = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]
        return jsonify({"success": True, "sites": sites})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching Site IDs and Site Names: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_area_ids', methods=['GET'])
def get_area_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Fetch Area IDs and Locations
        cursor.execute("SELECT `Area ID`, `Location` FROM `areas`")
        areas = [{"id": row[0], "location": row[1]} for row in cursor.fetchall()]
        return jsonify({"success": True, "areas": areas})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching Area IDs and Locations: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route("/get_row_ids", methods=["GET"])
def get_row_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Fetch Row IDs and Row Names from the database
        cursor.execute("SELECT `Row ID`, `Row Name` FROM `rows`")
        rows = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]
        return jsonify({"success": True, "rows": rows})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching rows: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route("/get_table_ids", methods=["GET"])
def get_table_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Fetch Table IDs and Locations from the database
        cursor.execute("SELECT `Table ID`, `Location` FROM `tables`")
        tables = [{"id": table[0], "location": table[1]} for table in cursor.fetchall()]
        return jsonify({"success": True, "tables": tables})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching tables: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_customer_ids', methods=['GET'])
def get_customer_ids():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Fetch Customer ID and Customer Name from the database
        cursor.execute("SELECT `Cust ID`, `Customer Name` FROM `customer`")
        customers = [{"id": row[0], "name": row[1]} for row in cursor.fetchall()]
        
        # Return the list of customers as a JSON response
        return jsonify({"success": True, "customers": customers})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching Customer IDs and Names: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_item_names', methods=['GET'])
def get_item_names():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        query = "SELECT `Item ID`, `Item Type` FROM `inventory`"
        cursor.execute(query)
        items = cursor.fetchall()

        item_list = [{"item_id": item[0], "item_name": item[1]} for item in items]

        return jsonify({"success": True, "items": item_list})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching item names: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_itemtrns_names', methods=['GET'])
def get_itemtrns_names():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        query = "SELECT `Item ID`, `Item Type` FROM `invtrans`"
        cursor.execute(query)
        items = cursor.fetchall()

        item_list = [{"item_id": item[0], "item_name": item[1]} for item in items]

        return jsonify({"success": True, "items": item_list})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching item names: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/search', methods=['GET'])
def search():
    query = request.args.get('query', '')  # Get the search query from the request
    results = []

    try:
        # Connect to the database using get_db_connection
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)

        # SQL Query to search for assessment names
        sql_query = """
        SELECT `assessment id`
        FROM assessment
        WHERE `assessment id` LIKE %s
        LIMIT 10
        """ 
        cursor.execute(sql_query, (f"%{query}%",))

        # Fetch all results
        results = cursor.fetchall()

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return jsonify({"error": str(err)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

    return jsonify({"results": results})


@app.route('/search_by_date', methods=['GET'])
def search_by_date():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    user_id = request.args.get('user_id')
    task_date = request.args.get('date')

    if not user_id or not task_date:
        return jsonify({"error": "User ID and Date parameters are required"}), 400

    try:
        # Fetch Table ID, Assessment Status, Assessment Case, and Pile No separately
        cursor.execute("""
            SELECT `Table ID`, `Assessment Status`, `Assessment Case`, `Pile No`
            FROM assessment
            WHERE `User ID` = %s 
            AND DATE(`Task Date`) = %s  
            
            ORDER BY `Table ID`, `Pile No`
        """, (user_id, task_date))

        result = cursor.fetchall()

        formatted_data = []
        for row in result:
            table_id = row["Table ID"]
            pile_no = row["Pile No"]

            formatted_data.append({
                "Assessment Case": row["Assessment Case"],
                "Assessment Status": row["Assessment Status"],
                "Pile No": pile_no,
                "Table ID": f"{table_id}-Pile{pile_no}"  # Format Table ID with Pile No
            })

        return jsonify({"data": formatted_data}) if formatted_data else jsonify({"data": []})

    except mysql.connector.Error as e:
        return jsonify({"error": f"Database error: {e}"}), 500

    finally:
        cursor.close()
        connection.close()




############################################################################

@app.route('/submit_updateuser_form', methods=['POST'])
def submit_updateuser_form():
    user_id = request.form.get('user_id')  
    user_email = request.form.get('user_email')  
    user_password = request.form.get('user_password')  
    phone_no = request.form.get('phone_no')
    date_removed = request.form.get('date_removed')
    user_type = request.form.get('user_type')  # Capture the user type from the form

    # Connect to the database
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        update_fields = []
        params = []

        if user_email:
            update_fields.append("`Email` = %s")
            params.append(user_email)

        if user_password:
            update_fields.append("`Password` = %s")
            params.append(user_password)

        if phone_no:
            update_fields.append("`User Phone number` = %s")
            params.append(phone_no)

        if date_removed:
            update_fields.append("`Date removed` = %s")
            params.append(date_removed)

        if user_type:  # Add a condition for user_type
            update_fields.append("`User Type` = %s")
            params.append(user_type)

        if not update_fields:
            return jsonify({"success": False, "message": "No fields to update provided."})

        params.append(user_id)
        query = f"""
            UPDATE `users`
            SET {', '.join(update_fields)}
            WHERE `User ID` = %s
        """
        cursor.execute(query, tuple(params))
        connection.commit()

        return jsonify({"success": True, "message": "User Details updated successfully."})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error updating user: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/delete_user/<user_id>', methods=['DELETE'])
def delete_user(user_id):
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Check if the user exists
        cursor.execute("SELECT `User ID` FROM `users` WHERE `User ID` = %s", (user_id,))
        if not cursor.fetchone():
            return jsonify({"success": False, "message": "User not found."})

        # Delete the user
        cursor.execute("DELETE FROM `users` WHERE `User ID` = %s", (user_id,))
        connection.commit()

        return jsonify({"success": True, "message": "User deleted successfully."})

    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error deleting user: {e}"})
    finally:
        cursor.close()
        connection.close()

# Update site details
@app.route("/submit_site_update", methods=["POST"])
def submit_site_update():
    try:
        data = request.form
        site_id = data.get("site_id")
        site_location = data.get("location")
        site_owner_name = data.get("site_owner_name")

        if not site_id or not site_location or not site_owner_name:
            return jsonify({"success": False, "message": "All fields are required."})

        connection = get_db_connection()
        cursor = connection.cursor()

        # Enclose column names with spaces in backticks
        query = """
            UPDATE `site` 
            SET `Site Location` = %s, `Site Owner Name` = %s 
            WHERE `Site ID` = %s
        """
        cursor.execute(query, (site_location, site_owner_name, site_id))
        connection.commit()

        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Site updated successfully."})
        else:
            return jsonify({"success": False, "message": "No changes made or site not found."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        cursor.close()
        connection.close()

# Delete a site
@app.route("/delete_site/<site_id>", methods=["DELETE"])
def delete_site(site_id):
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Use backticks for column names
        query = "DELETE FROM `site` WHERE `Site ID` = %s"
        cursor.execute(query, (site_id,))
        connection.commit()

        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Site deleted successfully."})
        else:
            return jsonify({"success": False, "message": "Site not found."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        cursor.close()
        connection.close()

# Route to handle customer update
@app.route('/submit_customer_update', methods=['POST'])
def submit_customer_update():
    try:
        # Get form data
        customer_id = request.form.get('customer_name')
        address = request.form.get('address')
        phone_no = request.form.get('phone_no')

        if not customer_id:
            return jsonify({"success": False, "message": "Customer ID is required."})

        # Initialize list to store the updates
        updates = []
        values = []

        # Only add fields if they are provided
        if address:
            updates.append("`Customer Address` = %s")
            values.append(address)
        if phone_no:
            updates.append("`Phone No` = %s")
            values.append(phone_no)

        # If no updates are provided, return an error
        if not updates:
            return jsonify({"success": False, "message": "At least one field should be provided to update."})

        # Add the customer_id to the end of values
        values.append(customer_id)

        # Create the SQL query dynamically based on provided fields
        query = f"""
            UPDATE `customer`
            SET {', '.join(updates)}
            WHERE `Cust ID` = %s
        """

        # Get database connection
        connection = get_db_connection()
        cursor = connection.cursor()

        # Execute the update query
        cursor.execute(query, tuple(values))
        connection.commit()

        # Check if the update was successful
        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Customer updated successfully."})
        else:
            return jsonify({"success": False, "message": "No changes made or customer not found."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        # Close the cursor and connection
        cursor.close()
        connection.close()

# Route to handle customer deletion
@app.route("/delete_customer/<customer_id>", methods=["DELETE"])
def delete_customer(customer_id):
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # Delete query with backticks for columns
        query = "DELETE FROM `customer` WHERE `Cust ID` = %s"
        cursor.execute(query, (customer_id,))
        connection.commit()

        if cursor.rowcount > 0:
            return jsonify({"success": True, "message": "Customer deleted successfully."})
        else:
            return jsonify({"success": False, "message": "Customer not found."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_customer_details/<customer_id>', methods=['GET'])
def get_customer_details(customer_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
       
        query = """
        SELECT `Customer Address`, `Phone No`
        FROM `Customer`
        WHERE `Cust ID` = %s
        """
        cursor.execute(query, (customer_id,))
        customer = cursor.fetchone()

        if customer:
           
            return jsonify({"success": True, "customer": {
                "address": customer[0],
                "phone_no": customer[1]
            }})
        else:
            return jsonify({"success": False, "message": "Customer not found."})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching customer details: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_inventory_details/<item_id>', methods=['GET'])
def get_inventory_details(item_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        query = """
        SELECT `Item Type`, `Item UOM`, `Item Avl Qty`, `Item ROR`, `Item Value`, `Item Rate`
        FROM `inventory`
        WHERE `Item ID` = %s
        """
        cursor.execute(query, (item_id,))
        item = cursor.fetchone()

        if item:
            return jsonify({
                "success": True,
                "inventory": {
                    "item_type": item[0],
                    "item_uom": item[1],
                    "item_avl_qty": item[2],
                    "item_ror": item[3],
                    "item_value": item[4],
                    "item_rate": item[5]
                }
            })
        else:
            return jsonify({"success": False, "message": "Item not found."})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching item details: {e}"})
    finally:
        cursor.close()
        connection.close()


@app.route('/get_site_details/<site_id>', methods=['GET'])
def get_site_details(site_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Query to fetch site details (location, site owner name)
        query = """
        SELECT `Site Location`, `Site Owner Name`
        FROM `Site`
        WHERE `Site ID` = %s
        """
        cursor.execute(query, (site_id,))
        site = cursor.fetchone()

        if site:
            # Return site details as JSON
            return jsonify({"success": True, "site": {
                "location": site[0],
                "owner_name": site[1]
            }})
        else:
            return jsonify({"success": False, "message": "Site not found."})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching site details: {e}"})
    finally:
        cursor.close()
        connection.close()

@app.route('/get_user_details/<user_id>', methods=['GET'])
def get_user_details(user_id):
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # Query to fetch user details (email, password, phone number, etc.)
        query = """
        SELECT `Email`, `Password`, `User Phone number`
        FROM `users`
        WHERE `User ID` = %s
        """
        cursor.execute(query, (user_id,))
        user = cursor.fetchone()

        if user:
            # Return user details as JSON
            return jsonify({
                "success": True,
                "user": {
                    "email": user[0],
                    "password": user[1],
                    "phone_no": user[2]
                }
            })
        else:
            return jsonify({"success": False, "message": "User not found."})
    except mysql.connector.Error as e:
        return jsonify({"success": False, "message": f"Error fetching user details: {e}"})
    finally:
        cursor.close()
        connection.close()
@app.route("/update_inventory_details", methods=["POST"])
def update_inventory_details():
    try:
        data = request.form
        item_id = data.get("item_type")  # Should be item_id, not item_type
        if not item_id:
            return jsonify({"success": False, "message": "Item ID is required for updating."})

        column_mapping = {
            "item_uom": "Item UOM",
            "item_desc": "Item Desc",
            "item_avl_qty": "Item Avl Qty",
            "item_ror": "Item ROR",
            "item_value": "Item Value",
            "item_rate": "Item Rate"
        }

        updates = []
        values = []

        for form_field, db_column in column_mapping.items():
            value = data.get(form_field)
            if value:
                updates.append(f"`{db_column}` = %s")
                values.append(value)

        if not updates:
            return jsonify({"success": True, "message": "No fields provided for update. Item is unchanged."})

        query = f"""
            UPDATE `inventory`
            SET {', '.join(updates)}
            WHERE `Item ID` = %s
        """
        values.append(item_id)  # Use Item ID as the identifier

        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(query, tuple(values))
        connection.commit()

        # Always return a success message
        return jsonify({"success": True, "message": "Inventory updated successfully."})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        if "cursor" in locals():
            cursor.close()
        if "connection" in locals():
            connection.close()



@app.route('/delete_item/<item_id>', methods=['DELETE'])
def delete_item(item_id):
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        # Execute delete query using Item ID
        cursor.execute("DELETE FROM inventory WHERE `Item ID` = %s", (item_id,))
        connection.commit()

        if cursor.rowcount > 0:
            return jsonify({
                'success': True,
                'message': f'Item with ID "{item_id}" deleted successfully.'
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Item with ID "{item_id}" not found.'
            }), 404
    except Exception as e:
        connection.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    finally:
        cursor.close()
        connection.close()

@app.route("/update_item_transaction_form", methods=["POST"])
def update_item_transaction_form():
    try:
        data = request.form
        item_id = data.get("item_type")  # Should be item_id, not item_name
        if not item_id:
            return jsonify({"success": False, "message": "Item ID is required for updating."})

        column_mapping = {
            "trans_qty": "Trans Qty",
            "trans_type": "Trans Type",
            "trans_date": "Trans Date",
            "user_id": "User ID",
            "usage": "Usage"
        }

        updates = []
        values = []

        for form_field, db_column in column_mapping.items():
            value = data.get(form_field)
            if value:
                updates.append(f"`{db_column}` = %s")
                values.append(value)

        if not updates:
            return jsonify({"success": True, "message": "No fields provided for update. Item is unchanged."})

        query = f"""
            UPDATE `invtrans`
            SET {', '.join(updates)}
            WHERE `Item ID` = %s
        """
        values.append(item_id)

        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(query, tuple(values))
        connection.commit()

        return jsonify({"success": True, "message": "Item updated successfully."})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        if "cursor" in locals():
            cursor.close()
        if "connection" in locals():
            connection.close()
            
#invtrns
@app.route('/delete_itemtrns/<item_id>', methods=['DELETE'])
def delete_itemtrns(item_id):
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("DELETE FROM invtrans WHERE `Item ID` = %s", (item_id,))
        connection.commit()

        return jsonify({
            'success': True,
            'message': f'Item with ID "{item_id}" deleted successfully.'
        })
    except Exception as e:
        connection.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
    finally:
        cursor.close()
        connection.close()

@app.route("/get_item_details/<item_id>", methods=["GET"])
def get_item_details(item_id):
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("SELECT * FROM invtrans WHERE `Item ID` = %s", (item_id,))
        item = cursor.fetchone()

        if item:
            return jsonify({"success": True, "item": item})
        else:
            return jsonify({"success": False, "message": "Item not found."})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        cursor.close()
        connection.close()

#############################################################################
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True, threaded=True)
